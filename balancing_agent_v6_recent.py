#!/usr/bin/env python3
"""
============================================================================
BALANCING REPORT GENERATOR - AI Agent v6.0
============================================================================
ROOT CAUSE FIX: Parses MISC-DATA-4 (not MISC-DATA-3) for cycle codes.
Activity Data: MISC-DATA-4 replaced by parsed sub-fields.
MISC-DATA-3 shown as numbers. Subtotals show "XX Total" format.

Usage:
1. Place BALDATA.txt, PRICDATA.txt, FLAPDATA.txt in one folder
2. Run: python balancing_agent_v6.py
3. Enter the folder path when prompted
Requirements: pip install pandas openpyxl numpy
============================================================================
"""
import pandas as pd
import numpy as np
import os, sys, zipfile
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

BAL_FIELDS = [
    "SPA", "PORTFOLIO", "EXT STATUS", "CYCLE CODE",
    "CURRENT-BALANCE", "OVERPAYMENT-AMT",
    "AP-OPEN-CYC-CASH-PRIN", "AP-OPEN-CYC-MRCH-NBINT",
    "AP-CURR-CTD-CASH-PRIN", "AP-CURR-CTD-MRCH-PRIN",
    "AP-CASH-MUF", "AP-MRCH-MUF",
    "AP-MISCELLANEOUS", "AP-OPEN-CYC-MRCH-BINT",
    "AP-MISC-CHGS", "AP-OPEN-CYC-FLAP-PRIN",
    "AP-OPEN-CYC-INTSC", "AP-MUF-CASH-INTSC", "AP-MUF-MRCH-INTSC",
    "AP-DISPUTED-AMOUNT", "AP-MUF-DISPUTE",
    "AP-UNPD-ANNL-CHRG-AM", "AP-UNPD-LATE-CHRG-AM",
    "AP-UNPD-CASH-ITEM-AM", "AP-UNPD-SALE-ITEM-AM",
    "AP-UNPD-OVRL-CHRG-AM", "AP-UNPD-CRDLF-PRMM-AM",
    "AP-MUF-UNPD-OVRL-CH-AM", "AP-MUF-UNPD-MSCL-CH-AM",
    "AP-MUF-UNPD-ANNL-CH-AM", "AP-MUF-UNPD-SRCH-AM",
    "AP-MUF-UNPD-LATE-CH-AM", "AP-MUF-UNPD-ITEM-CH-AM",
    "AP-MUF-UNPD-CRDT-LI-AM",
    "AP-TOTL-MMB-AM",
    "MUF-CTD-OVRL-CHRG-AM", "MUF-CTD-MSCL-CHRG-AM",
    "MUF-CTD-ANNL-CHRG-AM", "MUF-CTD-SRCH-AM",
    "MUF-CTD-LATE-CHRG-AM", "MUF-CTD-ITEM-CHRG-AM",
    "CTD-AMT-LATE-CHG", "CTD-OVERLIMIT-CHG",
    "CTD-AMT-ITEM-CHG", "CTD-AMT-MISC-CHGS",
    "CTD-SALE-ITEM-CHGS", "CTD-AMT-DEBITS",
    "CTD-ANNUAL-CHARGE",
    "NO-DEL-1CYC", "NO-DEL-2CYC", "NO-DEL-3CYC",
    "NO-DEL-4CYC", "NO-DEL-5CYC", "NO-DEL-6CYC", "NO-DEL-7CYC",
    "DEL-1CYC-AMT", "DEL-2CYC-AMT", "DEL-3CYC-AMT",
    "DEL-4CYC-AMT", "DEL-5CYC-AMT", "DEL-6CYC-AMT", "DEL-7CYC-AMT",
    "TOT-AMT-CURR-1CYC", "TOT-AMT-CURR-2CYC", "TOT-AMT-CURR-3CYC",
    "TOT-AMT-CURR-4CYC", "TOT-AMT-CURR-5CYC", "TOT-AMT-CURR-6CYC",
    "TOT-AMT-CURR-7CYC",
    "CREDIT-LINE", "HIST-LS-BAL", "CTD-UNPAID-BPD", "TOT-NO-OF-ACCTS"
]
FLAP_FIELDS = [
    "SPA", "PROMOTIONAL-ID", "BLNC-TYPE-CD",
    "ANN-INT-RATE", "ORIG-ANN-INT-RATE", "MTHD-RELD-CD",
    "CRRN-BLNC-AM", "CURR-UNPAID-INTSC", "BINT",
    "CTD-PRIN", "OLD-PRIN",
    "CTD-ITEM-FEES", "CTD-LATE-CHRG-AM", "CTD-OVRL-CHRG-AM",
    "CTD-MSCL-CHRG-AM", "CTD-ANNL-CHRG-AM", "CTD-BLLD-INTR-AM",
    "DISPUTE-AMOUNT",
    "UNPD-LATE-CHRG-AM", "UNPD-OVRL-CHRG-AM", "UNPD-CRDT-LIFE-AM",
    "UNPD-MSCL-CHRG-AM", "UNPD-ANNL-CHRG-AM", "UNPD-SRCHG-AM",
    "UNPD-ITEM-CHRG-AM",
    "CRDT-BLNC-AM", "TOT-NO-OF-FLAPS"
]
PRIC_FIELDS = [
    "SPA", "EXTENAL-STATUS",
    "CURR-PRICING-STRATEGY", "CURR-PRICING-PORTFOLIO",
    "CRRN-OVRR-TYPE-ID(1)", "CRRN-OVRR-ID(1)",
    "CRRN-OVRR-TYPE-ID(2)", "CRRN-OVRR-ID(2)",
    "CRRN-OVRR-TYPE-ID(3)", "CRRN-OVRR-ID(3)",
    "CRRN-OVRR-TYPE-ID(4)", "CRRN-OVRR-ID(4)",
    "CRRN-OVRR-TYPE-ID(5)", "CRRN-OVRR-ID(5)",
    "CRRN-OVRR-TYPE-ID(6)", "CRRN-OVRR-ID(6)",
    "CRRN-OVRR-TYPE-ID(7)", "CRRN-OVRR-ID(7)",
    "CRRN-OVRR-TYPE-ID(8)", "CRRN-OVRR-ID(8)",
    "CRRN-OVRR-TYPE-ID(9)", "CRRN-OVRR-ID(9)",
    "CRRN-OVRR-TYPE-ID(10)", "CRRN-OVRR-ID(10)",
    "CRRN-OVRR-TYPE-ID(11)", "CRRN-OVRR-ID(11)",
    "CRRN-OVRR-TYPE-ID(12)", "CRRN-OVRR-ID(12)",
    "CRRN-OVRR-TYPE-ID(13)", "CRRN-OVRR-ID(13)",
    "CRRN-OVRR-TYPE-ID(14)", "CRRN-OVRR-ID(14)",
    "CRRN-OVRR-TYPE-ID(15)", "CRRN-OVRR-ID(15)",
    "MISC-DATA-1", "MISC-DATA-2", "MISC-DATA-3", "MISC-DATA-4",
    "CURRENT-BALANCE", "TOT-NO-OF-ACCTS"
]

def to_num(s): return pd.to_numeric(s, errors='coerce').fillna(0)

def find_file(fp, bn):
    for n in [bn, bn.upper(), bn.lower(), bn.capitalize()]:
        for ext in ['.txt', '.txt.zip']:
            p = os.path.join(fp, n + ext)
            if os.path.exists(p): return p, ext.endswith('.zip')
    return None, None

def load_data_file(fp, zipped, flds):
    if fp is None: return None
    try:
        if zipped:
            with zipfile.ZipFile(fp,'r') as z:
                tn=[n for n in z.namelist() if n.lower().endswith('.txt')]
                if not tn: return None
                with z.open(tn[0]) as f:
                    df=pd.read_csv(f,sep='\t',header=None,dtype=str,keep_default_na=False)
        else:
            df=pd.read_csv(fp,sep='\t',header=None,dtype=str,keep_default_na=False)
        if len(df.columns)>len(flds): df=df.iloc[:,:len(flds)]
        elif len(df.columns)<len(flds):
            for i in range(len(df.columns),len(flds)): df[i]=''
        df.columns=flds
        return df
    except Exception as e:
        print(f"  ERROR loading {fp}: {e}"); return None

def add_grand_total(df, gcols):
    t={}
    for c in df.columns:
        if c==gcols[0]: t[c]='Grand Total'
        elif c in gcols[1:]: t[c]=''
        else: t[c]=df[c].sum() if pd.api.types.is_numeric_dtype(df[c]) else ''
    return pd.concat([df,pd.DataFrame([t])],ignore_index=True)

def add_computed_columns(bal_df):
    tc=['SPA','PORTFOLIO','EXT STATUS','CYCLE CODE']
    for c in bal_df.columns:
        if c not in tc: bal_df[c]=to_num(bal_df[c])
    bal_df['Total Prin']=(bal_df['AP-OPEN-CYC-FLAP-PRIN']+bal_df['AP-OPEN-CYC-MRCH-BINT']+bal_df['AP-OPEN-CYC-CASH-PRIN']+bal_df['AP-OPEN-CYC-MRCH-NBINT']+bal_df['AP-CURR-CTD-CASH-PRIN']+bal_df['AP-CURR-CTD-MRCH-PRIN']+bal_df['AP-CASH-MUF']+bal_df['AP-MRCH-MUF'])
    bal_df['Total Interest']=(bal_df['AP-OPEN-CYC-INTSC']+bal_df['AP-MUF-CASH-INTSC']+bal_df['AP-MUF-MRCH-INTSC'])
    bal_df['Total Fee']=(bal_df['AP-UNPD-ANNL-CHRG-AM']+bal_df['AP-UNPD-LATE-CHRG-AM']+bal_df['AP-UNPD-CASH-ITEM-AM']+bal_df['AP-UNPD-SALE-ITEM-AM']+bal_df['AP-UNPD-OVRL-CHRG-AM']+bal_df['AP-UNPD-CRDLF-PRMM-AM']+bal_df['AP-MUF-UNPD-OVRL-CH-AM']+bal_df['AP-MUF-UNPD-ANNL-CH-AM']+bal_df['AP-MUF-UNPD-SRCH-AM']+bal_df['AP-MUF-UNPD-LATE-CH-AM']+bal_df['AP-MUF-UNPD-ITEM-CH-AM']+bal_df['AP-MUF-UNPD-CRDT-LI-AM']+bal_df['MUF-CTD-OVRL-CHRG-AM']+bal_df['MUF-CTD-ANNL-CHRG-AM']+bal_df['MUF-CTD-SRCH-AM']+bal_df['MUF-CTD-LATE-CHRG-AM']+bal_df['MUF-CTD-ITEM-CHRG-AM']+bal_df['CTD-AMT-LATE-CHG']+bal_df['CTD-OVERLIMIT-CHG']+bal_df['CTD-AMT-ITEM-CHG']+bal_df['CTD-SALE-ITEM-CHGS']+bal_df['CTD-ANNUAL-CHARGE'])
    bal_df['Total Miscellaneous']=(bal_df['AP-MISCELLANEOUS']+bal_df['AP-MISC-CHGS']+bal_df['AP-MUF-UNPD-MSCL-CH-AM']+bal_df['AP-TOTL-MMB-AM']+bal_df['MUF-CTD-MSCL-CHRG-AM']+bal_df['CTD-AMT-MISC-CHGS']+bal_df['CTD-AMT-DEBITS'])
    bal_df['Total Dispute']=(bal_df['AP-DISPUTED-AMOUNT']+bal_df['AP-MUF-DISPUTE'])
    base=list(BAL_FIELDS); ip=base.index('AP-OPEN-CYC-FLAP-PRIN')+1
    comp=['Total Prin','Total Interest','Total Fee','Total Miscellaneous','Total Dispute']
    bal_df=bal_df[base[:ip]+comp+base[ip:]]
    return bal_df

def parse_misc_data4_row(raw_val):
    s = str(raw_val) if pd.notna(raw_val) else ''
    s = s.ljust(10)
    cc = s[0:2].strip()
    rfm = s[2:3].strip()
    gross = 'X' if s[3:4] == 'X' else ''
    debit = 'X' if s[4:5] == 'X' else ''
    active = 'X' if s[5:6] == 'X' else ''
    trash = s[6:].rstrip()
    return cc, rfm, gross, debit, active, trash

def parse_misc_data4_df(pric_df):
    if pric_df is None or pric_df.empty: return pric_df
    parsed = pric_df['MISC-DATA-4'].apply(parse_misc_data4_row)
    pric_df['CYCLE-CODE'] = parsed.apply(lambda x: x[0])
    pric_df['REC-FMT']    = parsed.apply(lambda x: x[1])
    pric_df['GROSS']       = parsed.apply(lambda x: x[2])
    pric_df['DEBIT']       = parsed.apply(lambda x: x[3])
    pric_df['ACTIVE']      = parsed.apply(lambda x: x[4])
    pric_df['TRASH']       = parsed.apply(lambda x: x[5])
    return pric_df

def build_summary_spa(b):
    d=b.groupby('SPA').agg(**{'Sum of TOT-NO-OF-ACCTS':('TOT-NO-OF-ACCTS','sum'),'Sum of CURRENT-BALANCE':('CURRENT-BALANCE','sum')}).reset_index()
    return add_grand_total(d,['SPA'])

def build_summary_ext(b):
    d=b.groupby(['SPA','EXT STATUS']).agg(**{'Sum of TOT-NO-OF-ACCTS':('TOT-NO-OF-ACCTS','sum'),'Sum of CURRENT-BALANCE':('CURRENT-BALANCE','sum')}).reset_index()
    return add_grand_total(d,['SPA','EXT STATUS'])

def build_balmstr(b):
    d=b.groupby(['SPA','PORTFOLIO']).agg(**{'Sum of CURRENT-BALANCE':('CURRENT-BALANCE','sum'),'Sum of Total Prin':('Total Prin','sum'),'Sum of Total Interest':('Total Interest','sum'),'Sum of Total Fee':('Total Fee','sum'),'Sum of Total Miscellaneous':('Total Miscellaneous','sum'),'Sum of Total Dispute':('Total Dispute','sum'),'Sum of OVERPAYMENT-AMT':('OVERPAYMENT-AMT','sum'),'Sum of TOT-NO-OF-ACCTS':('TOT-NO-OF-ACCTS','sum')}).reset_index()
    return add_grand_total(d,['SPA','PORTFOLIO'])

def build_flapmstr(f):
    if f is None or f.empty: return None
    for c in f.columns:
        if c not in ['SPA','PROMOTIONAL-ID','BLNC-TYPE-CD','MTHD-RELD-CD']: f[c]=to_num(f[c])
    d=f.groupby(['SPA','PROMOTIONAL-ID','ANN-INT-RATE','ORIG-ANN-INT-RATE','BLNC-TYPE-CD']).agg(**{'Sum of CRRN-BLNC-AM':('CRRN-BLNC-AM','sum'),'Sum of TOT-NO-OF-FLAPS':('TOT-NO-OF-FLAPS','sum')}).reset_index()
    return add_grand_total(d,['SPA','PROMOTIONAL-ID','ANN-INT-RATE','ORIG-ANN-INT-RATE','BLNC-TYPE-CD'])

def build_pricing_mstr(p):
    if p is None or p.empty: return None
    for c in ['CURRENT-BALANCE','TOT-NO-OF-ACCTS']: p[c]=to_num(p[c])
    d=p.groupby(['SPA','CURR-PRICING-STRATEGY','CURR-PRICING-PORTFOLIO']).agg(**{'Sum of CURRENT-BALANCE':('CURRENT-BALANCE','sum'),'Sum of TOT-NO-OF-ACCTS':('TOT-NO-OF-ACCTS','sum')}).reset_index()
    return add_grand_total(d,['SPA','CURR-PRICING-STRATEGY','CURR-PRICING-PORTFOLIO'])

def build_delqinfo(b):
    dc=[f'NO-DEL-{i}CYC' for i in range(1,8)]; da=[f'DEL-{i}CYC-AMT' for i in range(1,8)]
    ag={}
    for c in dc+da: ag[f'Sum of {c}']=(c,'sum')
    d=b.groupby('SPA').agg(**ag).reset_index()
    tc=sum(b[c].sum() for c in dc); ta=sum(b[c].sum() for c in da)
    return add_grand_total(d,['SPA']),tc,ta

def build_delqbalanceinfo(b):
    tc=[f'TOT-AMT-CURR-{i}CYC' for i in range(1,8)]
    ag={}
    for c in tc: ag[f'Sum of {c}']=(c,'sum')
    d=b.groupby('SPA').agg(**ag).reset_index()
    tb=sum(b[c].sum() for c in tc)
    return add_grand_total(d,['SPA']),tb

def build_extstatus(b):
    d=b.groupby('EXT STATUS').agg(**{'Sum of TOT-NO-OF-ACCTS':('TOT-NO-OF-ACCTS','sum')}).reset_index()
    return add_grand_total(d,['EXT STATUS'])

def build_cyclecodes(b):
    t=b.copy(); t['CYCLE CODE']=t['CYCLE CODE'].astype(str).str.strip().str.zfill(2)
    d=t.groupby('CYCLE CODE').agg(**{'Sum of TOT-NO-OF-ACCTS':('TOT-NO-OF-ACCTS','sum')}).reset_index()
    return add_grand_total(d,['CYCLE CODE'])

def build_activity_tables(pric_df):
    if pric_df is None or pric_df.empty: return None, None
    df=pric_df.copy()
    df['CURRENT-BALANCE']=to_num(df['CURRENT-BALANCE']); df['TOT-NO-OF-ACCTS']=to_num(df['TOT-NO-OF-ACCTS'])
    df=parse_misc_data4_df(df)
    v=df[df['CYCLE-CODE'].str.strip()!=''].copy()
    v=v[v['CYCLE-CODE'].str.isnumeric()].copy()
    v['CYCLE-CODE']=v['CYCLE-CODE'].str.zfill(2)
    vf=v[v['REC-FMT'].str.strip()!=''].copy()
    if vf.empty: cycle_df=None
    else:
        det=vf.groupby(['CYCLE-CODE','REC-FMT']).agg(**{'Sum of TOT-NO-OF-ACCTS':('TOT-NO-OF-ACCTS','sum')}).reset_index()
        rows=[]
        for cc in sorted(det['CYCLE-CODE'].unique()):
            cd=det[det['CYCLE-CODE']==cc].sort_values('REC-FMT')
            for _,r in cd.iterrows():
                rows.append({'CYCLE-CODE':r['CYCLE-CODE'],'REC-FMT':r['REC-FMT'],'Sum of TOT-NO-OF-ACCTS':r['Sum of TOT-NO-OF-ACCTS']})
            # SUBTOTAL: "XX Total" format matching macro output
            rows.append({'CYCLE-CODE':f'{cc} Total','REC-FMT':'','Sum of TOT-NO-OF-ACCTS':cd['Sum of TOT-NO-OF-ACCTS'].sum()})
        cycle_df=pd.DataFrame(rows)
        cycle_df=pd.concat([cycle_df,pd.DataFrame([{'CYCLE-CODE':'Grand Total','REC-FMT':'','Sum of TOT-NO-OF-ACCTS':det['Sum of TOT-NO-OF-ACCTS'].sum()}])],ignore_index=True)
    df['GROSS ACTIVE']=df.apply(lambda r:'X' if r.get('GROSS','')=='X' or r.get('DEBIT','')=='X' or r.get('ACTIVE','')=='X' else '',axis=1)
    gr=df.groupby('GROSS ACTIVE').agg(**{'Sum of CURRENT-BALANCE':('CURRENT-BALANCE','sum'),'Sum of TOT-NO-OF-ACCTS':('TOT-NO-OF-ACCTS','sum')}).reset_index()
    gr['GROSS ACTIVE']=gr['GROSS ACTIVE'].replace('','(blank)')
    gross_df=add_grand_total(gr,['GROSS ACTIVE'])
    return cycle_df, gross_df

def build_activity_data(pric_df):
    if pric_df is None or pric_df.empty: return None
    df=pric_df.copy()
    df['MISC-DATA-3'] = pd.to_numeric(df['MISC-DATA-3'], errors='coerce')
    df['CURRENT-BALANCE']=to_num(df['CURRENT-BALANCE']); df['TOT-NO-OF-ACCTS']=to_num(df['TOT-NO-OF-ACCTS'])
    df=parse_misc_data4_df(df)
    keep=[]; parsed=['CYCLE-CODE','REC-FMT','GROSS','DEBIT','ACTIVE','TRASH']
    for c in PRIC_FIELDS:
        if c == 'MISC-DATA-4':
            continue
        keep.append(c)
        if c == 'MISC-DATA-3':
            keep.extend(parsed)
    return df[[c for c in keep if c in df.columns]]

COUNT_KW=['ACCTS','FLAPS','NO-DEL']
MONEY_KW=['BALANCE','AMT','Prin','Interest','Fee','Miscellaneous','Dispute','OVERPAYMENT','BLNC','CREDIT-LINE','HIST-LS-BAL','BPD','INTSC','BINT','CHRG','SRCH','CRDT','ITEM']
CENTER_COLS=['SPA','EXT STATUS','EXTENAL-STATUS','CYCLE CODE','CYCLE-CODE','PORTFOLIO','BLNC-TYPE-CD','PROMOTIONAL-ID','REC-FMT','GROSS','DEBIT','ACTIVE','GROSS ACTIVE','MTHD-RELD-CD','CURR-PRICING-STRATEGY','CURR-PRICING-PORTFOLIO']
SPA_COLS=['SPA']
TAB_COLORS={"Directory":"70AD47","Summary":"00B050","BALMSTR":"4472C4","FLAPMSTR":"4472C4","PRICING MSTR":"4472C4","DelqInfo":"FF0000","DelqBalanceInfo":"FF0000","ExtStatus":"FFC000","CycleCodes":"FFC000","Activity":"00B050","Bal Data":"D9E2F3","Flap Data":"D9E2F3","Pricing Data":"D9E2F3","Activity Data":"D9E2F3","Field Names":"BDD7EE","Field Names 2":"BDD7EE","Field Names 3":"BDD7EE","Cross Check":"70AD47"}
HF=Font(name='Calibri',bold=True,size=10,color="FFFFFF")
HFL=PatternFill(start_color="4472C4",end_color="4472C4",fill_type="solid")
HA=Alignment(horizontal='center',wrap_text=True)
DF=Font(name='Calibri',size=10)
GF=Font(name='Calibri',bold=True,size=10)
GFL=PatternFill(start_color="D9E2F3",end_color="D9E2F3",fill_type="solid")
AF=PatternFill(start_color="F2F2F2",end_color="F2F2F2",fill_type="solid")
GRF=PatternFill(start_color="C6EFCE",end_color="C6EFCE",fill_type="solid")
SF=PatternFill(start_color="E2EFDA",end_color="E2EFDA",fill_type="solid")
TB=Border(left=Side(style='thin',color='D9D9D9'),right=Side(style='thin',color='D9D9D9'),top=Side(style='thin',color='D9D9D9'),bottom=Side(style='thin',color='D9D9D9'))
FF=Font(name='Calibri',bold=True,size=10,color="4472C4")

def is_cc(h): return any(k in str(h) for k in COUNT_KW)
def is_mc(h): return not is_cc(h) and any(k in str(h) for k in MONEY_KW)

def wcv(cell,v,h,cc,mc,spa=False):
    if pd.isna(v) or v=='': cell.value=''
    elif spa and str(v)!='Grand Total':
        try: cell.value=int(float(v)); cell.number_format='0'
        except: cell.value=v
    elif cc and not isinstance(v,str):
        try: cell.value=int(float(v)); cell.number_format='#,##0'
        except: cell.value=v
    elif mc and not isinstance(v,str):
        try: cell.value=float(v); cell.number_format='$#,##0.00'
        except: cell.value=v
    else: cell.value=v

def asc(ws,sc,ec,hr,dr,mw=30):
    for c in range(sc,ec+1):
        ml=10
        for r in range(hr,min(hr+dr+1,hr+102)):
            v=ws.cell(row=r,column=c).value
            if v is not None: ml=max(ml,len(str(v))+2)
        ws.column_dimensions[get_column_letter(c)].width=min(ml,mw)

def write_report_sheet(wb,name,df,tc=None,fl=None):
    if df is None or df.empty: return
    ws=wb.create_sheet(title=name)
    if tc: ws.sheet_properties.tabColor=tc
    hd=list(df.columns); nc=len(hd)
    ccc=[is_cc(h) for h in hd]; cmc=[is_mc(h) for h in hd]
    ccn=[h in CENTER_COLS for h in hd]; csp=[h in SPA_COLS for h in hd]
    cr=1
    if fl:
        for l,v in fl: ws.cell(row=cr,column=1,value=f"{l}: {v}").font=FF; cr+=1
        cr+=1
    hr=cr
    for ci,h in enumerate(hd,1):
        c=ws.cell(row=hr,column=ci,value=h); c.font=HF; c.fill=HFL; c.alignment=HA; c.border=TB
    for ri,(_,row) in enumerate(df.iterrows()):
        er=hr+1+ri; gt=str(row.iloc[0])=='Grand Total'; alt=(ri%2==0) and not gt
        for ci,(h,v) in enumerate(zip(hd,row),1):
            c=ws.cell(row=er,column=ci); wcv(c,v,h,ccc[ci-1],cmc[ci-1],csp[ci-1])
            c.font=GF if gt else DF
            if gt: c.fill=GFL
            elif alt: c.fill=AF
            if ccn[ci-1]: c.alignment=Alignment(horizontal='center')
            c.border=TB
    ws.auto_filter.ref=f"A{hr}:{get_column_letter(nc)}{hr+len(df)}"
    ws.freeze_panes=ws.cell(row=hr+1,column=1).coordinate
    asc(ws,1,nc,hr,len(df))

def write_summary_sheet(wb,spa_df,ext_df,tc=None):
    ws=wb.create_sheet(title="Summary")
    if tc: ws.sheet_properties.tabColor=tc
    def wt(df,sc,sr):
        hd=list(df.columns)
        for ci,h in enumerate(hd):
            c=ws.cell(row=sr,column=sc+ci,value=h); c.font=HF; c.fill=HFL; c.alignment=HA; c.border=TB
        for ri,(_,row) in enumerate(df.iterrows()):
            er=sr+1+ri; gt=str(row.iloc[0])=='Grand Total'
            for ci,(h,v) in enumerate(zip(hd,row)):
                c=ws.cell(row=er,column=sc+ci); wcv(c,v,h,is_cc(h),is_mc(h),h in SPA_COLS)
                c.font=GF if gt else DF
                if gt: c.fill=GFL
                elif ri%2==0: c.fill=AF
                if h in CENTER_COLS: c.alignment=Alignment(horizontal='center')
                c.border=TB
    wt(spa_df,1,1); wt(ext_df,len(spa_df.columns)+2,1)
    ws.freeze_panes="A2"
    for c in range(1,len(spa_df.columns)+len(ext_df.columns)+3):
        ws.column_dimensions[get_column_letter(c)].width=22

def write_activity_sheet(wb,cyc,grs,tc=None):
    ws=wb.create_sheet(title="Activity")
    if tc: ws.sheet_properties.tabColor=tc
    ws.cell(row=1,column=1,value="SPA: (All)").font=FF
    ws.cell(row=2,column=1,value="EXTENAL-STATUS: (All)").font=FF
    lr=3
    ws.cell(row=lr,column=1,value="Cycle Code and Record Format").font=Font(name='Calibri',bold=True,size=11,color="4472C4")
    hr=4
    if cyc is not None and not cyc.empty:
        th=list(cyc.columns)
        for ci,h in enumerate(th,1):
            c=ws.cell(row=hr,column=ci,value=h); c.font=HF; c.fill=HFL; c.alignment=HA; c.border=TB
        for ri,(_,row) in enumerate(cyc.iterrows()):
            er=hr+1+ri; gt=str(row.iloc[0])=='Grand Total'
            # Detect subtotal: "XX Total" rows
            sub='Total' in str(row.iloc[0]) and 'Grand' not in str(row.iloc[0])
            for ci,(h,v) in enumerate(zip(th,row),1):
                c=ws.cell(row=er,column=ci)
                if is_cc(h) and not isinstance(v,str):
                    try: c.value=int(float(v)); c.number_format='#,##0'
                    except: c.value=v
                else: c.value=v
                if gt: c.font=GF; c.fill=GFL
                elif sub: c.font=GF; c.fill=SF
                else: c.font=DF
                c.border=TB
    tc2=5
    ws.cell(row=lr,column=tc2,value="Gross Activity").font=Font(name='Calibri',bold=True,size=11,color="4472C4")
    if grs is not None and not grs.empty:
        gh=list(grs.columns)
        for ci,h in enumerate(gh):
            c=ws.cell(row=hr,column=tc2+ci,value=h); c.font=HF; c.fill=HFL; c.alignment=HA; c.border=TB
        for ri,(_,row) in enumerate(grs.iterrows()):
            er=hr+1+ri; gt=str(row.iloc[0])=='Grand Total'
            for ci,(h,v) in enumerate(zip(gh,row)):
                c=ws.cell(row=er,column=tc2+ci)
                if is_cc(h) and not isinstance(v,str):
                    try: c.value=int(float(v)); c.number_format='#,##0'
                    except: c.value=v
                elif is_mc(h) and not isinstance(v,str):
                    try: c.value=float(v); c.number_format='$#,##0.00'
                    except: c.value=v
                else: c.value=v
                c.font=GF if gt else DF
                if gt: c.fill=GFL
                c.border=TB
    for c in range(1,tc2+4): ws.column_dimensions[get_column_letter(c)].width=24

def write_delqinfo_sheet(wb,df,tc_val,ta,tcolor=None):
    ws=wb.create_sheet(title="DelqInfo")
    if tcolor: ws.sheet_properties.tabColor=tcolor
    ws.cell(row=1,column=1,value="EXT STATUS: (All)").font=FF
    ws.cell(row=2,column=1,value="Total Cnt Delq:").font=FF
    c2=ws.cell(row=2,column=2,value=int(tc_val)); c2.font=GF; c2.number_format='#,##0'
    ws.cell(row=3,column=1,value="Total Amt Delq:").font=FF
    c3=ws.cell(row=3,column=2,value=ta); c3.font=GF; c3.number_format='$#,##0.00'
    hr=5; hd=list(df.columns); nc=len(hd)
    ccc=[is_cc(h) for h in hd]; cmc=[is_mc(h) for h in hd]; csp=[h in SPA_COLS for h in hd]
    for ci,h in enumerate(hd,1):
        c=ws.cell(row=hr,column=ci,value=h); c.font=HF; c.fill=HFL; c.alignment=HA; c.border=TB
    for ri,(_,row) in enumerate(df.iterrows()):
        er=hr+1+ri; gt=str(row.iloc[0])=='Grand Total'
        for ci,(h,v) in enumerate(zip(hd,row),1):
            c=ws.cell(row=er,column=ci); wcv(c,v,h,ccc[ci-1],cmc[ci-1],csp[ci-1])
            c.font=GF if gt else DF
            if gt: c.fill=GFL
            elif ri%2==0: c.fill=AF
            c.border=TB
    ws.freeze_panes=f"A{hr+1}"; asc(ws,1,nc,hr,len(df))

def write_delqbalanceinfo_sheet(wb,df,tb,tcolor=None):
    ws=wb.create_sheet(title="DelqBalanceInfo")
    if tcolor: ws.sheet_properties.tabColor=tcolor
    ws.cell(row=1,column=1,value="Sum of Curr Bal on Delq Accts:").font=FF
    vc=ws.cell(row=1,column=2,value=tb); vc.number_format='$#,##0.00'; vc.font=GF
    ws.cell(row=2,column=1,value="EXT STATUS: (All)").font=FF
    hr=4; hd=list(df.columns); nc=len(hd)
    ccc=[is_cc(h) for h in hd]; cmc=[is_mc(h) for h in hd]; csp=[h in SPA_COLS for h in hd]
    for ci,h in enumerate(hd,1):
        c=ws.cell(row=hr,column=ci,value=h); c.font=HF; c.fill=HFL; c.alignment=HA; c.border=TB
    for ri,(_,row) in enumerate(df.iterrows()):
        er=hr+1+ri; gt=str(row.iloc[0])=='Grand Total'
        for ci,(h,v) in enumerate(zip(hd,row),1):
            c=ws.cell(row=er,column=ci); wcv(c,v,h,ccc[ci-1],cmc[ci-1],csp[ci-1])
            c.font=GF if gt else DF
            if gt: c.fill=GFL
            elif ri%2==0: c.fill=AF
            c.border=TB
    ws.freeze_panes=f"A{hr+1}"; asc(ws,1,nc,hr,len(df))

def write_raw_data_sheet(wb,name,df,tcolor=None):
    if df is None or df.empty: return
    ws=wb.create_sheet(title=name)
    if tcolor: ws.sheet_properties.tabColor=tcolor
    hd=list(df.columns); nc=len(hd); nr=len(df)
    for ci,h in enumerate(hd,1):
        c=ws.cell(row=1,column=ci,value=h); c.font=HF; c.fill=HFL; c.alignment=HA; c.border=TB
    for ri,(_,row) in enumerate(df.iterrows(),2):
        for ci,(h,v) in enumerate(zip(hd,row),1):
            c=ws.cell(row=ri,column=ci); spa=h in SPA_COLS
            if pd.isna(v) or v=='': c.value=''
            elif spa:
                try: c.value=int(float(v)); c.number_format='0'
                except: c.value=str(v)
            elif isinstance(v,(int,np.integer)): c.value=int(v)
            elif isinstance(v,(float,np.floating)): c.value=int(v) if v==int(v) else float(v)
            else:
                try:
                    fv=float(v); c.value=int(fv) if fv==int(fv) else fv
                except: c.value=str(v)
            c.font=DF
            if h in CENTER_COLS: c.alignment=Alignment(horizontal='center')
            c.border=TB
    if nr>0: ws.auto_filter.ref=f"A1:{get_column_letter(nc)}{nr+1}"
    ws.freeze_panes="A2"; asc(ws,1,nc,1,nr,25)

def write_directory_sheet(wb,fi,ts):
    ws=wb.active; ws.title="Directory"; ws.sheet_properties.tabColor=TAB_COLORS["Directory"]
    tf=Font(name='Calibri',bold=True,size=14,color="4472C4"); sf=Font(name='Calibri',bold=True,size=11); nf=Font(name='Calibri',size=10)
    ws.merge_cells('A1:D1'); ws['A1'].value="BALANCING REPORT - AI Agent v6.0"; ws['A1'].font=tf; ws['A1'].alignment=Alignment(horizontal='center')
    ws.merge_cells('A2:D2'); ws['A2'].value=f"Generated: {ts}"; ws['A2'].font=nf; ws['A2'].alignment=Alignment(horizontal='center')
    r=4; ws.cell(row=r,column=1,value="INPUT FILES").font=sf; r=5
    for ci,h in enumerate(["File Name","Status","Records Loaded"],1):
        c=ws.cell(row=r,column=ci,value=h); c.font=HF; c.fill=HFL; c.alignment=Alignment(horizontal='center'); c.border=TB
    for f in fi:
        r+=1; ws.cell(row=r,column=1,value=f['name']).font=nf
        sc=ws.cell(row=r,column=2,value=f['status']); sc.font=nf; sc.alignment=Alignment(horizontal='center')
        sc.fill=PatternFill(start_color="C6EFCE" if f['status']=='Loaded' else "FFC7CE",end_color="C6EFCE" if f['status']=='Loaded' else "FFC7CE",fill_type="solid")
        ws.cell(row=r,column=3,value=f['records']).font=nf; ws.cell(row=r,column=3).alignment=Alignment(horizontal='center')
        for ci in range(1,4): ws.cell(row=r,column=ci).border=TB
    r+=2; ws.cell(row=r,column=1,value="REPORT SHEETS").font=sf; r+=1
    for ci,h in enumerate(["Sheet Name","Description"],1):
        c=ws.cell(row=r,column=ci,value=h); c.font=HF; c.fill=HFL; c.alignment=Alignment(horizontal='center'); c.border=TB
    for sn,d in [("Summary","SPA summary + EXT STATUS breakdown"),("BALMSTR","Balance master by SPA, PORTFOLIO"),("FLAPMSTR","FLAP master"),("PRICING MSTR","Pricing by SPA, strategy, portfolio"),("DelqInfo","Delinquency counts and amounts"),("DelqBalanceInfo","Delinquency balance buckets"),("ExtStatus","External status account counts"),("CycleCodes","Cycle code account counts"),("Activity","Cycle Code x REC-FMT + Gross Activity"),("Bal Data","Raw BALDATA with computed columns"),("Flap Data","Raw FLAPDATA"),("Pricing Data","Raw PRICDATA"),("Activity Data","Pricing + parsed MISC-DATA-4 fields"),("Field Names","BALDATA field definitions"),("Field Names 2","FLAPDATA field definitions"),("Field Names 3","PRICDATA field definitions"),("Cross Check","Balance verification")]:
        r+=1; ws.cell(row=r,column=1,value=sn).font=nf; ws.cell(row=r,column=2,value=d).font=nf
        for ci in range(1,3): ws.cell(row=r,column=ci).border=TB
    ws.column_dimensions['A'].width=25; ws.column_dimensions['B'].width=80; ws.column_dimensions['C'].width=18

def write_field_names_sheet(wb,sn,fl,tc=None):
    ws=wb.create_sheet(title=sn)
    if tc: ws.sheet_properties.tabColor=tc
    for ci,h in enumerate(["Field #","Field Name"],1):
        c=ws.cell(row=1,column=ci,value=h); c.font=HF; c.fill=HFL; c.alignment=HA; c.border=TB
    for ri,fn in enumerate(fl,2):
        ws.cell(row=ri,column=1,value=ri-1).font=DF; ws.cell(row=ri,column=1).alignment=Alignment(horizontal='center'); ws.cell(row=ri,column=1).border=TB
        ws.cell(row=ri,column=2,value=fn).font=DF; ws.cell(row=ri,column=2).border=TB
    ws.freeze_panes="A2"; ws.column_dimensions['A'].width=12; ws.column_dimensions['B'].width=30

def write_cross_check_sheet(wb,bal_df,tc=None):
    ws=wb.create_sheet(title="Cross Check")
    if tc: ws.sheet_properties.tabColor=tc
    nf=Font(name='Calibri',size=10); sf=Font(name='Calibri',bold=True,size=10)
    ws.cell(row=1,column=1,value="BALANCE CROSS CHECK").font=Font(name='Calibri',bold=True,size=12,color="4472C4")
    ws.merge_cells('A1:C1'); ws['A1'].alignment=Alignment(horizontal='center')
    checks=[("CURRENT-BALANCE","Total Current Balance"),("TOT-NO-OF-ACCTS","Total Accounts"),("Total Prin","Total Principal"),("Total Interest","Total Interest"),("Total Fee","Total Fees"),("Total Miscellaneous","Total Miscellaneous"),("Total Dispute","Total Dispute"),("OVERPAYMENT-AMT","Total Overpayment"),("CREDIT-LINE","Total Credit Line")]
    r=3
    for ci,h in enumerate(["Measure","Column Name","Sum"],1):
        c=ws.cell(row=r,column=ci,value=h); c.font=HF; c.fill=HFL; c.alignment=Alignment(horizontal='center'); c.border=TB
    for cn,lb in checks:
        r+=1; ws.cell(row=r,column=1,value=lb).font=nf; ws.cell(row=r,column=1).border=TB
        ws.cell(row=r,column=2,value=cn).font=nf; ws.cell(row=r,column=2).border=TB
        if cn in bal_df.columns:
            cl=ws.cell(row=r,column=3,value=bal_df[cn].sum()); cl.number_format='#,##0' if 'ACCTS' in cn else '$#,##0.00'
        else: cl=ws.cell(row=r,column=3,value="N/A")
        cl.font=nf; cl.border=TB
    r+=2; ws.cell(row=r,column=1,value="BALANCE VERIFICATION").font=sf; r+=1
    tb=bal_df['CURRENT-BALANCE'].sum()
    cs=bal_df['Total Prin'].sum()+bal_df['Total Interest'].sum()+bal_df['Total Fee'].sum()+bal_df['Total Miscellaneous'].sum()+bal_df['Total Dispute'].sum()+bal_df['OVERPAYMENT-AMT'].sum()
    ws.cell(row=r,column=1,value="Sum of CURRENT-BALANCE:").font=nf; ws.cell(row=r,column=2,value=tb).font=nf; ws.cell(row=r,column=2).number_format='$#,##0.00'
    r+=1; ws.cell(row=r,column=1,value="Sum of Components:").font=nf; ws.cell(row=r,column=2,value=cs).font=nf; ws.cell(row=r,column=2).number_format='$#,##0.00'
    r+=1; diff=abs(tb-cs); ws.cell(row=r,column=1,value="Difference:").font=sf
    dc=ws.cell(row=r,column=2,value=diff); dc.number_format='$#,##0.00'
    if diff<0.01:
        dc.fill=GRF; dc.font=Font(name='Calibri',bold=True,size=10,color="006100")
        ws.cell(row=r,column=3,value="\u2713 BALANCE PROVES").font=Font(name='Calibri',bold=True,size=10,color="006100"); ws.cell(row=r,column=3).fill=GRF
    else:
        dc.fill=PatternFill(start_color="FFC7CE",end_color="FFC7CE",fill_type="solid"); dc.font=Font(name='Calibri',bold=True,size=10,color="9C0006")
        ws.cell(row=r,column=3,value="\u2717 VARIANCE DETECTED").font=Font(name='Calibri',bold=True,size=10,color="9C0006")
    ws.column_dimensions['A'].width=30; ws.column_dimensions['B'].width=25; ws.column_dimensions['C'].width=25

def generate_balancing_report(folder_path, output_filename=None):
    ts=datetime.now(); tss=ts.strftime("%Y-%m-%d %H:%M:%S"); tsf=ts.strftime("%Y%m%d_%H%M%S")
    if output_filename is None: output_filename=f"Balancing_Report_{tsf}.xlsx"
    op=os.path.join(folder_path,output_filename)
    print(f"\n  Report Generation Started: {tss}\n  Folder: {folder_path}\n  "+"-"*55)
    fi=[]
    bp,bz=find_file(folder_path,"BALDATA"); bd=load_data_file(bp,bz,BAL_FIELDS)
    if bd is not None: print(f"  \u2713 BALDATA.txt: {len(bd):,} records"); fi.append({'name':'BALDATA.txt','status':'Loaded','records':len(bd)})
    else: print("  \u2717 BALDATA.txt NOT FOUND"); fi.append({'name':'BALDATA.txt','status':'NOT FOUND','records':0}); return None
    fp,fz=find_file(folder_path,"FLAPDATA"); fd=load_data_file(fp,fz,FLAP_FIELDS)
    if fd is not None: print(f"  \u2713 FLAPDATA.txt: {len(fd):,} records"); fi.append({'name':'FLAPDATA.txt','status':'Loaded','records':len(fd)})
    else: print("  - FLAPDATA.txt not found"); fi.append({'name':'FLAPDATA.txt','status':'Not Found','records':0})
    pp,pz=find_file(folder_path,"PRICDATA"); pd_=load_data_file(pp,pz,PRIC_FIELDS)
    if pd_ is not None: print(f"  \u2713 PRICDATA.txt: {len(pd_):,} records"); fi.append({'name':'PRICDATA.txt','status':'Loaded','records':len(pd_)})
    else: print("  - PRICDATA.txt not found"); fi.append({'name':'PRICDATA.txt','status':'Not Found','records':0})
    print("  "+"-"*55)
    bd=add_computed_columns(bd)
    spv=bd['SPA'].unique(); spf=str(spv[0]) if len(spv)==1 else '(All)'
    print("  Building reports...")
    ss=build_summary_spa(bd); se=build_summary_ext(bd); print("    \u2713 Summary")
    bm=build_balmstr(bd); print("    \u2713 BALMSTR")
    fm=build_flapmstr(fd.copy() if fd is not None else None)
    if fm is not None: print("    \u2713 FLAPMSTR")
    pm=build_pricing_mstr(pd_.copy() if pd_ is not None else None)
    if pm is not None: print("    \u2713 PRICING MSTR")
    di,dc,da=build_delqinfo(bd); print("    \u2713 DelqInfo")
    db,tb=build_delqbalanceinfo(bd); print("    \u2713 DelqBalanceInfo")
    es=build_extstatus(bd); print("    \u2713 ExtStatus")
    cc=build_cyclecodes(bd); print("    \u2713 CycleCodes")
    ac,ag=build_activity_tables(pd_.copy() if pd_ is not None else None)
    if ac is not None: print("    \u2713 Activity")
    ad=build_activity_data(pd_.copy() if pd_ is not None else None)
    if ad is not None: print("    \u2713 Activity Data")
    print("\n  Writing Excel workbook...")
    wb=Workbook()
    write_directory_sheet(wb,fi,tss); print("    \u2713 Directory")
    write_summary_sheet(wb,ss,se,TAB_COLORS["Summary"]); print("    \u2713 Summary")
    write_report_sheet(wb,"BALMSTR",bm,TAB_COLORS["BALMSTR"],fl=[("EXT STATUS","(All)")]); print("    \u2713 BALMSTR")
    write_report_sheet(wb,"FLAPMSTR",fm,TAB_COLORS["FLAPMSTR"]); print("    \u2713 FLAPMSTR")
    write_report_sheet(wb,"PRICING MSTR",pm,TAB_COLORS["PRICING MSTR"],fl=[("EXTENAL-STATUS","(All)")]); print("    \u2713 PRICING MSTR")
    write_delqinfo_sheet(wb,di,dc,da,TAB_COLORS["DelqInfo"]); print("    \u2713 DelqInfo")
    write_delqbalanceinfo_sheet(wb,db,tb,TAB_COLORS["DelqBalanceInfo"]); print("    \u2713 DelqBalanceInfo")
    write_report_sheet(wb,"ExtStatus",es,TAB_COLORS["ExtStatus"],fl=[("SPA",spf)]); print("    \u2713 ExtStatus")
    write_report_sheet(wb,"CycleCodes",cc,TAB_COLORS["CycleCodes"],fl=[("SPA","(All)")]); print("    \u2713 CycleCodes")
    write_activity_sheet(wb,ac,ag,TAB_COLORS["Activity"]); print("    \u2713 Activity")
    write_raw_data_sheet(wb,"Bal Data",bd,TAB_COLORS["Bal Data"])
    write_raw_data_sheet(wb,"Flap Data",fd,TAB_COLORS["Flap Data"])
    write_raw_data_sheet(wb,"Pricing Data",pd_,TAB_COLORS["Pricing Data"])
    write_raw_data_sheet(wb,"Activity Data",ad,TAB_COLORS["Activity Data"]); print("    \u2713 Raw data sheets")
    write_field_names_sheet(wb,"Field Names",BAL_FIELDS,TAB_COLORS["Field Names"])
    write_field_names_sheet(wb,"Field Names 2",FLAP_FIELDS,TAB_COLORS["Field Names 2"])
    write_field_names_sheet(wb,"Field Names 3",PRIC_FIELDS,TAB_COLORS["Field Names 3"]); print("    \u2713 Field Names")
    write_cross_check_sheet(wb,bd,TAB_COLORS["Cross Check"]); print("    \u2713 Cross Check")
    wb.save(op)
    print(f"\n  \u2713 Report saved: {op}\n  \u2713 Sheets: {', '.join(wb.sheetnames)}")
    return op

if __name__ == "__main__":
    print("\n+" + "="*58 + "+")
    print("|  BALANCING REPORT AI AGENT v6.0                          |")
    print("|  ROOT CAUSE FIX: Parses MISC-DATA-4 for cycle codes      |")
    print("|  Subtotals: 'XX Total' format matching macro              |")
    print("+" + "="*58 + "+\n")
    print("Prerequisites: BALDATA.txt (required), PRICDATA.txt, FLAPDATA.txt (optional)\n")
    folder = input("Enter folder path containing data files: ").strip()
    if not folder: folder = "."
    if not os.path.isdir(folder): print(f"\n  ERROR: '{folder}' is not a valid directory."); sys.exit(1)
    generate_balancing_report(folder)
    print("\nDone!")
