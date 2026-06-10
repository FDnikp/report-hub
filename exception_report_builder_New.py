#!/usr/bin/env python3
"""Exception Report Builder v4.0"""
import os, re, sys, zipfile, tempfile, warnings
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from datetime import datetime
warnings.filterwarnings("ignore")

DESC_FILE_PATH = r"G:\zSecure1\conv\CONVPROJDOC\Conversions and Startups\DMG\Excel - Exception Template and Description File\Exception Description New.xlsx"
DESC_WORKSHEET = "Descriptions"

# Event to Event Summary styles
ETE_HDR_FILL = PatternFill(start_color="203864", end_color="203864", fill_type="solid")
ETE_HDR_FONT = Font(bold=True, color="FFFFFF", size=11)
ETE_LIGHT = PatternFill(start_color="B4C6E7", end_color="B4C6E7", fill_type="solid")
ETE_MED = PatternFill(start_color="8FAADC", end_color="8FAADC", fill_type="solid")
# All Exceptions styles
EXC_HDR_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
EXC_HDR_FONT = Font(bold=True, color="FFFFFF")
EXC_LIGHT = PatternFill(start_color="FDE9D9", end_color="FDE9D9", fill_type="solid")
EXC_ALT = PatternFill(start_color="8DB4E2", end_color="8DB4E2", fill_type="solid")
# Directory styles
DIR_BAN_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
DIR_BAN_FONT = Font(bold=True, color="FFFFFF", size=16)
DIR_SEC_FILL = PatternFill(start_color="203864", end_color="203864", fill_type="solid")
DIR_SEC_FONT = Font(bold=True, color="FFFFFF", size=12)
DIR_LBL_FILL = PatternFill(start_color="D6DCE4", end_color="D6DCE4", fill_type="solid")
DIR_LBL_FONT = Font(bold=True, size=11, color="1F4E79")
DIR_VAL_FONT = Font(size=11)
DIR_VAL_FILL = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
DIR_OK_FILL = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
DIR_OK_FONT = Font(bold=True, size=11, color="375623")
# Common
THIN_B = Border(left=Side(style="thin",color="A6A6A6"),right=Side(style="thin",color="A6A6A6"),top=Side(style="thin",color="A6A6A6"),bottom=Side(style="thin",color="A6A6A6"))
THICK_B = Border(left=Side(style="thick",color="1F4E79"),right=Side(style="thick",color="1F4E79"),top=Side(style="thin",color="A6A6A6"),bottom=Side(style="thin",color="A6A6A6"))
CTR = Alignment(horizontal="center", vertical="center")
WRAP = Alignment(wrap_text=True, vertical="top")
BOLD = Font(bold=True)
INT_FMT = "#,##0"
CHANGE_FMT = '#,##0;(#,##0);"-"'
COMMA_FMT = '_(* #,##0_);_(* (#,##0);_(* "-"??_);_(@_)'

def auto_fit(ws, mn=8, mx=56):
    for cc in ws.columns:
        lg = max((len(str(c.value)) for c in cc if c.value), default=0)
        ws.column_dimensions[get_column_letter(cc[0].column)].width = min(max(lg+2,mn),mx)

def extract_zip(zp):
    td = tempfile.mkdtemp()
    with zipfile.ZipFile(zp,"r") as zf:
        zf.extractall(td)
        return [os.path.join(td,n) for n in zf.namelist()]

def parse_sev(t):
    if not t or len(t)<2: return None,t
    if len(t)>=3 and t[1:3]==": ": return t[0],t[3:].strip()
    if t[1]==":": return t[0],t[2:].strip()
    return None,t

def find_file(folder, name):
    for f in os.listdir(folder):
        if f.lower()==name.lower(): return os.path.join(folder,f)
    return None

def find_exc(folder):
    for n in ["EXCPXLS.txt","EXCPRPT1.txt","EXCPXLS.zip","EXCPRPT1.zip"]:
        f=find_file(folder,n)
        if f: return f
    return None

def find_sum(folder):
    for n in ["EXCPSUM.txt","EXCPSUM.zip"]:
        f=find_file(folder,n)
        if f: return f
    return None

def normalize_desc(s):
    """Collapse multiple spaces to single space for matching."""
    return re.sub(r"\s+", " ", s.strip())

def clean_sheet_name(name):
    """Remove chars Excel doesnt allow in sheet names, truncate to 31."""
    for ch in [":", "\\", "/", "?", "*", "[", "]", "_"]:
        name = name.replace(ch, " ")
    return name[:31].strip()

class ExceptionReportBuilder:
    def __init__(self):
        self.folder_path=""
        self.exception_files=[]  # list of (filepath, filename, ws_name, summary_flag)
        self.events=[]
        self.desc_file=DESC_FILE_PATH
        self.summary_data=[]
        self.last_col=0
        self._pivot=None
        self._evts=[]
        self.desc_cols=0
        self.wb=openpyxl.Workbook()
        self.wb.remove(self.wb.active)

    def collect_inputs(self):
        print("\n"+"="*70)
        print("  EXCEPTION REPORT BUILDER v4.0")
        print("="*70)

        # STEP 1: Exception Folder Path
        print("\n  STEP 1: Exception Folder")
        print("  "+"-"*40)
        while True:
            p=input("  Enter exception folder path: ").strip().strip('"').strip("'")
            if not p: print("  [!] Empty path."); continue
            if not os.path.isdir(p): print(f"  [!] Not found: {p}"); continue
            self.folder_path=p
            print(f"  [OK] Folder: {self.folder_path}")
            break

        # Auto-detect default exception file
        default_file = find_exc(self.folder_path)
        if default_file:
            dfname = os.path.basename(default_file)
            print(f"  [OK] Auto-detected: {dfname}")
            self.exception_files = [(default_file, dfname, "All Exceptions", "Y")]
        else:
            print("  [!] No default exception file (EXCPXLS/EXCPRPT1) found.")
            txt_files = sorted([f for f in os.listdir(self.folder_path) if f.lower().endswith((".txt",".zip"))])
            if txt_files:
                print("  Available files:")
                for f in txt_files: print(f"       - {f}")
                fname = input("  Enter exception filename: ").strip()
                if fname:
                    fpath = os.path.join(self.folder_path, fname)
                    if os.path.isfile(fpath):
                        self.exception_files = [(fpath, fname, "All Exceptions", "Y")]
                        print(f"  [OK] Using: {fname}")
                    else:
                        print(f"  [!] File not found: {fname}")

        # STEP 1b: Additional Exception Files
        print("\n  STEP 1b: Additional Exception Files")
        print("  "+"-"*40)
        add_more = input("  Load additional exception files? (Y/N, default N): ").strip().upper()

        if add_more == "Y":
            # CUSTOMIZED MODE - clear default, user picks ALL files
            self.exception_files = []
            print("\n  [CUSTOMIZED MODE] Provide all exception file names.")
            print(f"  Files in: {self.folder_path}\n")
            txt_files = sorted([f for f in os.listdir(self.folder_path) if f.lower().endswith((".txt",".zip"))])
            for f in txt_files: print(f"       - {f}")
            print()
            file_num = 1
            while True:
                print(f"  --- Exception File #{file_num} ---")
                fname = input("  File name (or done): ").strip()
                if fname.lower() == "done":
                    if not self.exception_files:
                        print("  [!] Need at least one file.")
                        continue
                    break
                fpath = os.path.join(self.folder_path, fname)
                if not os.path.isfile(fpath):
                    print(f"  [!] Not found: {fname}")
                    continue
                ws_name = input(f"  Worksheet name for {fname} (Enter=use filename): ").strip()
                if not ws_name:
                    ws_name = clean_sheet_name(os.path.splitext(fname)[0])
                else:
                    ws_name = clean_sheet_name(ws_name)
                is_sum = input(f"  Is this a summary report? Y/N (default Y): ").strip().upper()
                if is_sum != "N": is_sum = "Y"
                self.exception_files.append((fpath, fname, ws_name, is_sum))
                print(f"  [OK] File #{file_num}: {fname} -> \"{ws_name}\"\n")
                file_num += 1
        else:
            if self.exception_files:
                print(f"  Using default: {self.exception_files[0][1]} -> \"{self.exception_files[0][2]}\"")
            else:
                print("  [!] No exception files configured.")

        print(f"\n  Total exception files: {len(self.exception_files)}")

        # STEP 2: Summary Events
        print("\n  STEP 2: Summary Events")
        print("  "+"-"*40)
        print("  Type done when finished.\n")
        n=1
        while True:
            print(f"  --- Event #{n} ---")
            nm=input(f"  Event name (e.g. Mock 1, LIVE) or done: ").strip()
            if nm.lower()=="done":
                if not self.events: print("  [!] Need at least 1 event."); continue
                break
            ep=input(f"  Folder for \"{nm}\": ").strip().strip('"').strip("'")
            if not ep or not os.path.isdir(ep): print(f"  [!] Not found: {ep}"); continue
            sf=find_sum(ep)
            if not sf:
                fs=sorted([f for f in os.listdir(ep) if f.lower().endswith((".txt",".zip"))])
                print(f"  [!] EXCPSUM not found in: {ep}")
                if fs:
                    for f in fs[:15]: print(f"       - {f}")
                    c=input("  Enter summary filename: ").strip()
                    if c: sf=os.path.join(ep,c)
                if not sf or not os.path.isfile(sf): continue
            self.events.append((nm,ep,sf))
            print(f"  [OK] Event #{n}: {nm}\n"); n+=1

        # STEP 3: Description File
        print("\n  STEP 3: Exception Description")
        print("  "+"-"*40)
        if os.path.isfile(self.desc_file):
            print(f"  [OK] Auto-loaded: {os.path.basename(self.desc_file)}")
        else:
            print(f"  [!] Not found: {self.desc_file}")
            alt=input("  Alternate path (Enter=skip): ").strip().strip('"').strip("'")
            if alt and os.path.isfile(alt): self.desc_file=alt
            else: self.desc_file=""
        print(f"\n  {'='*60}")
        print("  Building report...")
        print(f"  {'='*60}")

    def create_directory_sheet(self):
        print("\n  Creating Directory...")
        ws=self.wb.create_sheet(title="Directory")
        # Banner
        for c in range(1,5): ws.cell(row=1,column=c).fill=DIR_BAN_FILL
        ws.merge_cells("A1:D1")
        ws.cell(row=1,column=1,value="Exception Report Builder").font=DIR_BAN_FONT
        ws.cell(row=1,column=1).alignment=Alignment(horizontal="center",vertical="center")
        ws.row_dimensions[1].height=40
        # Subtitle
        for c in range(1,5): ws.cell(row=2,column=c).fill=PatternFill(start_color="2E75B6",end_color="2E75B6",fill_type="solid")
        ws.merge_cells("A2:D2")
        ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        ws.cell(row=2,column=1,value=f"Generated: {ts}").font=Font(color="FFFFFF",size=10,italic=True)
        ws.cell(row=2,column=1).alignment=Alignment(horizontal="center")
        # Exception Files section
        r=4
        for c in range(1,5): ws.cell(row=r,column=c).fill=DIR_SEC_FILL
        ws.merge_cells(f"A{r}:D{r}")
        ws.cell(row=r,column=1,value="  Exception Files Loaded").font=DIR_SEC_FONT
        r+=1
        ws.cell(row=r,column=1,value="Exception Folder:").font=DIR_LBL_FONT; ws.cell(row=r,column=1).fill=DIR_LBL_FILL; ws.cell(row=r,column=1).border=THIN_B
        ws.merge_cells(f"B{r}:D{r}")
        ws.cell(row=r,column=2,value=self.folder_path).font=DIR_VAL_FONT; ws.cell(row=r,column=2).fill=DIR_VAL_FILL; ws.cell(row=r,column=2).border=THIN_B
        r+=1
        for c,h in enumerate(["#","File Name","Worksheet Name","Summary Report"],start=1):
            ws.cell(row=r,column=c,value=h).font=DIR_LBL_FONT; ws.cell(row=r,column=c).fill=DIR_LBL_FILL; ws.cell(row=r,column=c).border=THIN_B
        for i,(fp,fn,wn,sf) in enumerate(self.exception_files,start=1):
            r+=1
            ws.cell(row=r,column=1,value=i).font=DIR_VAL_FONT; ws.cell(row=r,column=1).border=THIN_B; ws.cell(row=r,column=1).alignment=CTR
            ws.cell(row=r,column=2,value=fn).font=Font(bold=True,size=11); ws.cell(row=r,column=2).border=THIN_B
            ws.cell(row=r,column=3,value=wn).font=DIR_VAL_FONT; ws.cell(row=r,column=3).border=THIN_B
            ws.cell(row=r,column=4,value=sf).font=DIR_OK_FONT; ws.cell(row=r,column=4).fill=DIR_OK_FILL; ws.cell(row=r,column=4).border=THIN_B; ws.cell(row=r,column=4).alignment=CTR
        # Events section
        r+=2
        for c in range(1,5): ws.cell(row=r,column=c).fill=DIR_SEC_FILL
        ws.merge_cells(f"A{r}:D{r}")
        ws.cell(row=r,column=1,value="  Summary Events Used").font=DIR_SEC_FONT
        r+=1
        for c,h in enumerate(["#","Event Name","Summary File Path","Status"],start=1):
            ws.cell(row=r,column=c,value=h).font=DIR_LBL_FONT; ws.cell(row=r,column=c).fill=DIR_LBL_FILL; ws.cell(row=r,column=c).border=THIN_B
        for i,(en,ep,ef) in enumerate(self.events,start=1):
            r+=1
            ws.cell(row=r,column=1,value=i).font=DIR_VAL_FONT; ws.cell(row=r,column=1).border=THIN_B; ws.cell(row=r,column=1).alignment=CTR
            ws.cell(row=r,column=2,value=en).font=Font(bold=True,size=11); ws.cell(row=r,column=2).border=THIN_B
            ws.cell(row=r,column=3,value=ef).font=DIR_VAL_FONT; ws.cell(row=r,column=3).border=THIN_B
            ws.cell(row=r,column=4,value="Loaded").font=DIR_OK_FONT; ws.cell(row=r,column=4).fill=DIR_OK_FILL; ws.cell(row=r,column=4).border=THIN_B; ws.cell(row=r,column=4).alignment=CTR
        # Description section
        r+=2
        for c in range(1,5): ws.cell(row=r,column=c).fill=DIR_SEC_FILL
        ws.merge_cells(f"A{r}:D{r}")
        ws.cell(row=r,column=1,value="  Description File Information").font=DIR_SEC_FONT
        for label,val in [("Description File:",self.desc_file or "(None)"),("Worksheet Name:",DESC_WORKSHEET if self.desc_file else "(N/A)"),("Add to Event-to-Event:","Y")]:
            r+=1
            ws.cell(row=r,column=1,value=label).font=DIR_LBL_FONT; ws.cell(row=r,column=1).fill=DIR_LBL_FILL; ws.cell(row=r,column=1).border=THIN_B
            ws.cell(row=r,column=2,value=val).font=DIR_VAL_FONT; ws.cell(row=r,column=2).fill=DIR_VAL_FILL; ws.cell(row=r,column=2).border=THIN_B
        ws.column_dimensions["A"].width=28; ws.column_dimensions["B"].width=30; ws.column_dimensions["C"].width=55; ws.column_dimensions["D"].width=18
        print("    [OK] Directory")

    def load_exceptions(self):
        print("\n  Loading exception files...")
        if not self.exception_files:
            print("    No exception files."); return
        for idx,(filepath,fname,ws_name,sflag) in enumerate(self.exception_files,start=1):
            print(f"    [{idx}/{len(self.exception_files)}] {fname} -> \"{ws_name}\"", end="")
            af=filepath; tf=[]
            if filepath.lower().endswith(".zip"):
                if not os.path.isfile(filepath): print(" [SKIP]"); continue
                tf=extract_zip(filepath); af=tf[0] if tf else None
                if not af: print(" [SKIP]"); continue
            elif not os.path.isfile(filepath): print(" [SKIP]"); continue
            try: df=pd.read_csv(af,sep="\t",dtype=str,keep_default_na=False)
            except Exception as e: print(f" [ERR:{e}]"); continue
            if df.empty: print(" [EMPTY]"); continue
            nr,nc=len(df),len(df.columns)
            final_name=ws_name
            counter=1
            while final_name in self.wb.sheetnames:
                final_name=f"{ws_name[:28]}_{counter}"; counter+=1
            ws=self.wb.create_sheet(title=final_name)
            for c,cn in enumerate(df.columns,start=1):
                cl=ws.cell(row=1,column=c,value=cn); cl.fill=EXC_HDR_FILL; cl.font=EXC_HDR_FONT; cl.border=THIN_B
            for r,(_,row) in enumerate(df.iterrows(),start=2):
                for c,v in enumerate(row,start=1):
                    cl=ws.cell(row=r,column=c,value=str(v) if v!="" else ""); cl.number_format="@"; cl.fill=EXC_LIGHT; cl.border=THIN_B
            ldr=nr+1; mc=nc
            if sflag=="Y" and nc>=8:
                h=8; j=10
                for cf in range(nc+1,j+1):
                    cl=ws.cell(row=1,column=cf,value=""); cl.fill=EXC_HDR_FILL; cl.font=EXC_HDR_FONT; cl.border=THIN_B
                for r in range(2,ldr+1):
                    hv=str(ws.cell(row=r,column=h).value or "")
                    sv,cl2=parse_sev(hv)
                    if sv:
                        ws.cell(row=r,column=h,value=cl2).number_format="@"
                        c2=ws.cell(row=r,column=j,value=sv); c2.number_format="@"; c2.fill=EXC_LIGHT; c2.border=THIN_B
                mc=max(nc,j)
                dr=[]
                for r in range(2,ldr+1): dr.append([ws.cell(row=r,column=c).value for c in range(1,mc+1)])
                def sk(rd):
                    d=str(rd[h-1] or "").lower()
                    try: cn=float(str(rd[2] or 0).replace(",",""))
                    except: cn=0
                    return(d,cn)
                dr.sort(key=sk)
                for r,rd in enumerate(dr,start=2):
                    for c,v in enumerate(rd,start=1): ws.cell(row=r,column=c,value=v).number_format="@"
                alt=False; pd2=str(ws.cell(row=2,column=h).value or "")
                for r in range(2,ldr+1):
                    cd=str(ws.cell(row=r,column=h).value or "")
                    if cd!=pd2: alt=not alt
                    fl=EXC_ALT if alt else EXC_LIGHT
                    for c in range(1,mc+1): ws.cell(row=r,column=c).fill=fl; ws.cell(row=r,column=c).border=THIN_B
                    pd2=cd
            for r in range(1,ldr+1):
                for c in [1,2,3,6,7]:
                    if c<=mc: ws.cell(row=r,column=c).alignment=CTR
            ws.auto_filter.ref=f"A1:{get_column_letter(mc)}{ldr}"
            ws.freeze_panes="B2"; auto_fit(ws)
            if mc>=10: ws.column_dimensions["J"].width=56
            print(f" [{nr} rows]")
            for t in tf:
                try: os.remove(t)
                except: pass

    def build_exception_summary(self):
        print("\n  Building Exception Summary...")
        if not self.events: return
        ar=[]
        for i,(en,ep,ef) in enumerate(self.events,start=1):
            ns=str(i) if i>=10 else f"0{i}"
            lb=f"({ns}) {en}"
            af=ef; tf=[]
            if ef.lower().endswith(".zip"):
                if not os.path.isfile(ef): continue
                tf=extract_zip(ef); af=tf[0] if tf else None
                if not af: continue
            elif not os.path.isfile(ef): continue
            try: df=pd.read_csv(af,sep="\t",dtype=str,keep_default_na=False,header=None)
            except: continue
            for _,row in df.iterrows():
                d=str(row.iloc[0]).strip() if len(row)>0 else ""
                cs=str(row.iloc[1]).strip() if len(row)>1 else ""
                d=d.replace("** OMITTED **","Z- OMITTED --")
                d=normalize_desc(d)  # FIX: collapse extra spaces
                try: cv=int(float(cs.replace(",","")))
                except: cv=0
                ar.append([d,cv,lb,"",""])
            print(f"    [OK] {lb} ({len(df)} types)")
            for t in tf:
                try: os.remove(t)
                except: pass
        if not ar: return
        self.summary_data=ar
        for row in self.summary_data:
            sv,cl=parse_sev(row[0])
            if sv: row[3]=sv; row[0]=cl
        sm={}
        for row in self.summary_data:
            if row[3]: sm[row[0]]=row[3]
        for row in self.summary_data:
            if row[0] in sm: row[4]=sm[row[0]]
        ws=self.wb.create_sheet(title="Exception Summary Work")
        for c,h in enumerate(["Description","Count","Event","Event Severity","Current Severity"],start=1): ws.cell(row=1,column=c,value=h).font=BOLD
        for r,rd in enumerate(self.summary_data,start=2):
            for c,v in enumerate(rd,start=1): ws.cell(row=r,column=c,value=v)
        ws.auto_filter.ref=f"A1:E{len(self.summary_data)+1}"; auto_fit(ws); ws.sheet_state="hidden"
        print(f"    [OK] Summary Work: {len(self.summary_data)} rows")

    def build_event_to_event_summary(self):
        print("\n  Building Event-to-Event Summary...")
        if not self.summary_data: self.wb.create_sheet(title="Event to Event Summary",index=0); return
        df=pd.DataFrame(self.summary_data,columns=["Description","Count","Event","ES","CS"])
        df["Count"]=pd.to_numeric(df["Count"],errors="coerce").fillna(0).astype(int)
        df=df[df["Description"].str.strip()!=""]
        if df.empty: self.wb.create_sheet(title="Event to Event Summary",index=0); return
        pv=df.pivot_table(index="Description",columns="Event",values="Count",aggfunc="sum",fill_value=0)
        pv=pv.reindex(sorted(pv.columns),axis=1).reset_index().sort_values("Description").reset_index(drop=True)
        ws=self.wb.create_sheet(title="Event to Event Summary",index=0)
        ec=[c for c in pv.columns if c!="Description"]
        ne=len(ec)
        # Row 1
        ws.cell(row=1,column=1,value="Sum of Count").font=BOLD
        ws.cell(row=1,column=2,value="Column Labels")
        # Row 2 headers with COLORING
        c2=ws.cell(row=2,column=1,value="Exception Label")
        c2.font=ETE_HDR_FONT; c2.fill=ETE_HDR_FILL; c2.alignment=Alignment(wrap_text=True,vertical="center"); c2.border=THIN_B
        for ci,ev in enumerate(ec,start=2):
            cl=ws.cell(row=2,column=ci,value=ev)
            cl.font=ETE_HDR_FONT; cl.fill=ETE_HDR_FILL; cl.alignment=Alignment(wrap_text=True,vertical="center",horizontal="center"); cl.border=THIN_B
        ws.row_dimensions[2].height=35
        # Data rows with alternating colors
        for ri,(_,row) in enumerate(pv.iterrows(),start=3):
            fill=ETE_LIGHT if (ri%2==1) else ETE_MED
            cl=ws.cell(row=ri,column=1,value=row["Description"])
            cl.fill=fill; cl.border=THIN_B; cl.font=Font(size=10)
            for ci,ev in enumerate(ec,start=2):
                v=int(row[ev])
                cl=ws.cell(row=ri,column=ci,value=v if v!=0 else None)
                cl.number_format=CHANGE_FMT; cl.fill=fill; cl.border=THIN_B; cl.alignment=CTR
                if v==0: cl.value=None; cl.number_format="General"
        ldr=len(pv)+2; lc=1+ne; self.last_col=lc
        # Change column
        cc=None
        if ne>1:
            cc=lc+1
            ch=ws.cell(row=2,column=cc,value="Change Since Prior Event")
            ch.font=ETE_HDR_FONT; ch.fill=ETE_HDR_FILL; ch.alignment=Alignment(wrap_text=True,vertical="center",horizontal="center"); ch.border=THICK_B
            for r in range(3,ldr+1):
                lv=ws.cell(row=r,column=lc).value or 0
                pv2=ws.cell(row=r,column=lc-1).value or 0
                d=int(lv)-int(pv2)
                fill=ETE_LIGHT if (r%2==1) else ETE_MED
                cl=ws.cell(row=r,column=cc,value=d if d!=0 else None)
                cl.number_format=CHANGE_FMT; cl.border=THICK_B; cl.fill=fill; cl.alignment=CTR
                if d==0: cl.value=None; cl.number_format="General"
            self.last_col=cc
        ec2=cc or lc
        ws.auto_filter.ref=f"A2:{get_column_letter(ec2)}{ldr}"
        auto_fit(ws); ws.column_dimensions["A"].width=52
        self._pivot=pv; self._evts=ec
        print(f"    [OK] Pivot: {len(pv)} exceptions x {ne} events")

    def load_description(self):
        print("\n  Loading Description...")
        if not self.desc_file or not os.path.isfile(self.desc_file): print("    Skipping."); return
        try: ddf=pd.read_excel(self.desc_file,sheet_name=DESC_WORKSHEET,dtype=str,engine="openpyxl").fillna("")
        except Exception as e: print(f"    [!] {e}"); return
        if len(ddf.columns)>0: ddf.iloc[:,0]=ddf.iloc[:,0].str.strip()
        self.desc_cols=len(ddf.columns)
        dcn=list(ddf.columns)
        # Description Work (hidden)
        ww=self.wb.create_sheet(title="Description Work")
        for c,cn in enumerate(ddf.columns,start=1): ww.cell(row=1,column=c,value=cn)
        for r,(_,row) in enumerate(ddf.iterrows(),start=2):
            for c,v in enumerate(row,start=1): ww.cell(row=r,column=c,value=v)
        ww.sheet_state="hidden"
        # Lookup
        dl={}
        for _,row in ddf.iterrows(): dl[str(row.iloc[0]).strip()]=list(row.iloc[1:])
        # Get exceptions from last event
        dr=[]
        if self._pivot is not None and len(self._evts)>0:
            le=self._evts[-1]
            for _,row in self._pivot.iterrows():
                nm=str(row["Description"]).strip(); ct=int(row[le])
                if ct==0 or nm.startswith("Z-"): continue
                dr.append((nm,ct))
        # Description sheet
        wd=self.wb.create_sheet(title="Description")
        wd.cell(row=1,column=1,value="Exception"); wd.cell(row=1,column=2,value="Count")
        for c,cn in enumerate(dcn[1:],start=3): wd.cell(row=1,column=c,value=cn)
        for r,(nm,ct) in enumerate(dr,start=2):
            wd.cell(row=r,column=1,value=nm); wd.cell(row=r,column=2,value=ct).number_format=COMMA_FMT
            if nm in dl:
                for c,v in enumerate(dl[nm],start=3): wd.cell(row=r,column=c,value=v)
        ldr2=max(len(dr)+1,2); tc=max(2,self.desc_cols+1)
        if dr:
            tr=f"A1:{get_column_letter(tc)}{ldr2}"
            tb=Table(displayName="Table1",ref=tr)
            tb.tableStyleInfo=TableStyleInfo(name="TableStyleMedium9",showFirstColumn=False,showLastColumn=False,showRowStripes=True,showColumnStripes=False)
            wd.add_table(tb)
        for r in range(1,ldr2+1):
            for c in range(3,tc+1): wd.cell(row=r,column=c).number_format="@"; wd.cell(row=r,column=c).alignment=WRAP
        auto_fit(wd); wd.freeze_panes="A2"
        print(f"    [OK] Description: {len(dr)} exceptions")
        # Add to Event-to-Event
        if "Event to Event Summary" in self.wb.sheetnames:
            we=self.wb["Event to Event Summary"]; sc=self.last_col+1
            for d,cn in enumerate(dcn[1:]):
                cp=sc+d; cl=we.cell(row=2,column=cp,value=cn)
                cl.font=ETE_HDR_FONT; cl.fill=ETE_HDR_FILL; cl.alignment=Alignment(wrap_text=True,vertical="center"); cl.border=THIN_B
            for r in range(3,we.max_row+1):
                nm=str(we.cell(row=r,column=1).value or "").strip()
                fill=ETE_LIGHT if (r%2==1) else ETE_MED
                if nm in dl:
                    for d,v in enumerate(dl[nm]):
                        cp=sc+d; cl=we.cell(row=r,column=cp,value=v); cl.number_format="@"; cl.alignment=WRAP; cl.fill=fill; cl.border=THIN_B
                else:
                    for d in range(len(dcn)-1):
                        cp=sc+d; cl=we.cell(row=r,column=cp); cl.fill=fill; cl.border=THIN_B
            for d in range(len(dcn)-1): we.column_dimensions[get_column_letter(sc+d)].width=50
            print(f"    [OK] Added {len(dcn)-1} desc columns to ETE")

    def save_output(self):
        print("\n  Saving output...")
        # Build desired sheet order: Directory, ETE, exception sheets, Description
        priority=["Directory","Event to Event Summary"]
        for _,_,wn,_ in self.exception_files:
            if wn not in priority: priority.append(wn)
        priority.append("Description")
        for i,nm in enumerate(priority):
            if nm in self.wb.sheetnames:
                ws=self.wb[nm]; ci=self.wb.sheetnames.index(nm)
                if ci!=i: self.wb.move_sheet(ws,offset=i-ci)
        op=os.path.join(self.folder_path,"Exception_Report_Output.xlsx")
        try: self.wb.save(op); print(f"\n  {'='*60}"); print(f"  OUTPUT SAVED: {op}"); print(f"  {'='*60}")
        except PermissionError:
            alt=os.path.join(self.folder_path,"Exception_Report_Output_NEW.xlsx")
            self.wb.save(alt); print(f"  Saved as: {alt}")

    def run(self):
        self.collect_inputs()
        self.create_directory_sheet()
        self.load_exceptions()
        self.build_exception_summary()
        self.build_event_to_event_summary()
        self.load_description()
        self.save_output()
        print("\n  DONE!\n")

if __name__=="__main__":
    ExceptionReportBuilder().run()
