#!/usr/bin/env python3
"""
=============================================================================
  Report Tool  -  Standalone Python Report Generator  (v2)
=============================================================================
  Replaces the VBA Report Macro entirely.
  No macro file needed -- just point to a folder and go.

  FEATURES:
    - Auto-discovers .txt / .txt.zip report files
    - Only processes known report types by default (skip ACCT, BALDATA, etc.)
    - Smart column classification: Account Numbers stored as text (no
      scientific notation), Amount/Balance columns always show 2 decimals,
      Expiration fields zero-padded (425 -> 0425)
    - Auto-detects delimiters and column types
    - Professional Excel formatting: dark navy header, centered cells,
      freeze panes, auto-filter, auto-fit widths
    - Summary sheet and meaningful Data Profile per report
    - Modular design ready for AI Agent integration

  USAGE (Interactive):
    python report_tool.py

  USAGE (AI Agent -- no prompts):
    from report_tool import run_report_tool
    results = run_report_tool(r"G:\path\to\reports")

  REQUIREMENTS:
    pip install openpyxl
=============================================================================
"""

import os
import sys
import zipfile
import datetime
from collections import Counter

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter


# ===========================================================================
#  NAME MAPPING -- longer / more-specific patterns FIRST
# ===========================================================================
NAME_MAP = [
    ("CLOGRPTX", "CONV LOG"),
    ("CLOGRPT",  "CONV LOG"),
    ("CLOGX",    "CONV LOG"),
    ("CLOG",     "CONV LOG"),
    ("CROSSPI",  "CROSS REF PI REPORT"),
    ("CROSS",    "CROSS REF REPORT"),
    ("FRAUDPS",  "FRAUD PASSER"),
    ("FRAUD",    "FRAUD PASSER"),
    ("FAILX",    "FAIL"),
    ("FAIL",     "FAIL"),
    ("PASSX",    "PASS"),
    ("PASS",     "PASS"),
    ("STMTHOLD", "STMT HOLD PASSER"),
    ("STMTHLD",  "STMT HOLD PASSER"),
    ("NONTRUE",  "NONTRUE"),
    ("ESTMT",    "ESTMT"),
    ("NAMEPASR", "NAME PASSER"),
    ("CIFRPT",   "CIF REPORT"),
]

# Column classification codes
COL_TEXT       = "Text"
COL_NUMERIC    = "Numeric"
COL_ACCOUNT    = "Account Number"
COL_AMOUNT     = "Amount"
COL_EXPIRATION = "Expiration"

# ------- Header styling (dark navy) -------
HEADER_FILL = PatternFill(start_color="002060", end_color="002060",
                          fill_type="solid")
HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=12)
HEADER_ALIGN = Alignment(horizontal="center", vertical="center",
                         wrap_text=True)
HEADER_BORDER = Border(
    bottom=Side(style="medium", color="001040"),
    top=Side(style="thin", color="001040"),
    left=Side(style="thin", color="808080"),
    right=Side(style="thin", color="808080"),
)
DATA_ALIGN = Alignment(horizontal="center", vertical="center")

# Profile / Summary header colours
PROFILE_HEADER_FILL = PatternFill(start_color="00626F", end_color="00626F",
                                  fill_type="solid")
SUMMARY_HEADER_FILL = PatternFill(start_color="375623", end_color="375623",
                                  fill_type="solid")


# ===========================================================================
#  MODULE 1: FILE DISCOVERY
# ===========================================================================
def discover_files(folder_path):
    """Scan folder for .txt and .txt.zip files."""
    files = []
    if not os.path.isdir(folder_path):
        return files
    for fname in sorted(os.listdir(folder_path)):
        lower = fname.lower()
        if lower.endswith(".txt.zip") or lower.endswith(".txt"):
            fpath = os.path.join(folder_path, fname)
            if os.path.isfile(fpath):
                size_kb = round(os.path.getsize(fpath) / 1024, 2)
                ext = ".txt.zip" if lower.endswith(".txt.zip") else ".txt"
                files.append({
                    "name": fname,
                    "path": fpath,
                    "size_kb": size_kb,
                    "extension": ext,
                })
    return files


# ===========================================================================
#  MODULE 2: NAME MAPPING
# ===========================================================================
def get_file_stem(filename):
    """Remove .txt.zip or .txt extension, return uppercase stem."""
    lower = filename.lower()
    if lower.endswith(".txt.zip"):
        return filename[:-8]
    elif lower.endswith(".txt"):
        return filename[:-4]
    return filename


def map_filename(filename):
    """Map input filename to friendly output name (without .xlsx)."""
    stem = get_file_stem(filename).upper()
    for pattern, friendly in NAME_MAP:
        if stem == pattern:
            return friendly
    return stem


def _name_map_index(filename):
    """Return the index in NAME_MAP, or 9999 if not found."""
    stem = get_file_stem(filename).upper()
    for idx, (pattern, _) in enumerate(NAME_MAP):
        if stem == pattern:
            return idx
    return 9999


def select_default_files(all_files):
    """
    From all discovered files, select only those matching NAME_MAP.
    When duplicates map to the same output, pick the earlier NAME_MAP match.
    Returns (selected_files, other_files).
    """
    # Group by mapped output name
    groups = {}
    for finfo in all_files:
        friendly = map_filename(finfo["name"])
        nm_idx = _name_map_index(finfo["name"])
        if nm_idx == 9999:
            continue  # not in NAME_MAP
        if friendly not in groups:
            groups[friendly] = []
        groups[friendly].append((nm_idx, finfo))

    # For each group, pick the file with the lowest NAME_MAP index
    selected = []
    selected_names = set()
    for friendly in groups:
        candidates = sorted(groups[friendly], key=lambda x: x[0])
        best = candidates[0][1]
        selected.append(best)
        selected_names.add(best["name"])

    # Sort selected by original order
    selected.sort(key=lambda f: f["name"])

    others = [f for f in all_files if f["name"] not in selected_names]
    return selected, others


# ===========================================================================
#  MODULE 3: DELIMITER AUTO-DETECTION
# ===========================================================================
def detect_delimiter(filepath):
    """Auto-detect best delimiter from first 20 lines."""
    content = _read_raw_text(filepath)
    lines = [ln for ln in content.splitlines() if ln.strip()][:20]
    if not lines:
        return "|"

    candidates = ["|", ",", "\t", ";"]
    best_delim = "|"
    best_score = 0

    for delim in candidates:
        counts = [len(line.split(delim)) for line in lines]
        if not counts:
            continue
        most_common_count, freq = Counter(counts).most_common(1)[0]
        if most_common_count > 1:
            score = freq * most_common_count
            if score > best_score:
                best_score = score
                best_delim = delim
    return best_delim


# ===========================================================================
#  MODULE 4: FILE READING
# ===========================================================================
def _read_raw_text(filepath):
    """Read raw text from .txt or .txt.zip; UTF-8 with Latin-1 fallback."""
    if filepath.lower().endswith(".zip"):
        try:
            with zipfile.ZipFile(filepath, "r") as zf:
                names = zf.namelist()
                if not names:
                    return ""
                with zf.open(names[0]) as f:
                    raw = f.read()
        except zipfile.BadZipFile:
            return ""
    else:
        with open(filepath, "rb") as f:
            raw = f.read()

    for enc in ("utf-8", "latin-1"):
        try:
            return raw.decode(enc)
        except UnicodeDecodeError:
            continue
    return raw.decode("utf-8", errors="replace")


def parse_file(filepath, delimiter="|"):
    """Parse a delimited file into list of lists."""
    content = _read_raw_text(filepath)
    rows = []
    for line in content.splitlines():
        if line.strip() == "":
            continue
        parts = line.split(delimiter)
        rows.append([p.strip() for p in parts])
    return rows


# ===========================================================================
#  MODULE 5: SMART COLUMN CLASSIFICATION
# ===========================================================================
def _is_numeric_value(val):
    """Check if string can be interpreted as a number."""
    if val is None or val.strip() == "":
        return False
    clean = val.strip().replace(",", "")
    if clean.endswith("-") and not clean.startswith("-"):
        clean = "-" + clean[:-1]
    try:
        float(clean)
        return True
    except ValueError:
        return False


def _header_words(h):
    """Split header into uppercase words on _ - . and space."""
    result = h.upper()
    for ch in ["_", "-", ".", " "]:
        result = result.replace(ch, "|")
    return [w for w in result.split("|") if w]


def _is_account_header(h):
    """Check if header name indicates an account/card number column."""
    up = h.upper().replace(" ", "_")
    substrings = [
        "ACCOUNT_N", "ACCT_N", "CARD_NO", "CARD_NUM", "CARDNO",
        "ACCT_NO", "ACCT_NBR", "CARD_NBR", "MEMBER_NO", "PAN_NO",
    ]
    for s in substrings:
        if s in up:
            return True
    exact_words = {"ACCOUNT", "ACCT", "PAN", "CARD", "CID", "TOKEN"}
    words = _header_words(h)
    for w in words:
        if w in exact_words:
            return True
    return False


def _is_amount_header(h):
    """Check if header name indicates an amount/balance column."""
    up = h.upper().replace(" ", "_")
    substrings = [
        "AMOUNT", "BALANCE", "DELINQ", "INTEREST", "PAYMENT",
        "PURCHASE", "FINANCE", "OVERLIMIT", "CASH_ADV", "TOTAL_DUE",
        "PRINCIPAL", "ESCROW", "ANNUAL_FEE", "LATE_FEE", "CREDIT_LIM",
        "CREDIT_LINE", "CURR_BAL", "CUR_BAL", "NEW_BAL", "HIGH_BAL",
        "OPEN_BAL", "CLOSE_BAL", "BAL_AMT",
    ]
    for s in substrings:
        if s in up:
            return True
    exact_words = {"AMT", "BAL", "FEE", "DUE"}
    words = _header_words(h)
    for w in words:
        if w in exact_words:
            return True
    return False


def _is_expiration_header(h):
    """Check if header name indicates an expiration date column."""
    up = h.upper().replace(" ", "_")
    substrings = ["EXPIR", "XPDT", "EXP_DT", "EXP_DATE", "EXPDT"]
    for s in substrings:
        if s in up:
            return True
    exact_words = {"EXP"}
    words = _header_words(h)
    for w in words:
        if w in exact_words:
            return True
    return False


def classify_columns(rows):
    """
    Analyze header names and data patterns to classify each column.
    Returns list of classification strings.
    """
    if not rows:
        return []

    header = rows[0]
    data_rows = rows[1:]
    num_cols = len(header)
    sample_end = min(len(data_rows), 200)
    classifications = []

    for col_idx in range(num_cols):
        h = header[col_idx] if col_idx < len(header) else ""

        # Gather sample values
        values = []
        for ri in range(sample_end):
            if ri < len(data_rows) and col_idx < len(data_rows[ri]):
                values.append(data_rows[ri][col_idx].strip())
            else:
                values.append("")

        non_blank = [v for v in values if v != ""]
        total = len(non_blank)

        # 1. Check EXPIRATION first (header-based)
        if _is_expiration_header(h):
            classifications.append(COL_EXPIRATION)
            continue

        # 2. Check ACCOUNT NUMBER (header or data pattern)
        if _is_account_header(h):
            classifications.append(COL_ACCOUNT)
            continue
        # Data pattern: >50% have >10 digits => likely account number
        if total > 0:
            long_num_count = 0
            for v in non_blank:
                digits = v.replace(",", "").replace("-", "").replace(".", "")
                if digits.isdigit() and len(digits) > 10:
                    long_num_count += 1
            if (long_num_count / total) > 0.50:
                classifications.append(COL_ACCOUNT)
                continue

        # 3. Check AMOUNT (header or data pattern)
        numeric_count = sum(1 for v in non_blank if _is_numeric_value(v))
        has_decimal = any("." in v for v in non_blank if _is_numeric_value(v))

        if _is_amount_header(h):
            classifications.append(COL_AMOUNT)
            continue
        # Data pattern: >80% numeric AND some decimals
        if total > 0 and (numeric_count / total) > 0.80 and has_decimal:
            classifications.append(COL_AMOUNT)
            continue

        # 4. Regular NUMERIC
        if total > 0 and (numeric_count / total) > 0.80:
            classifications.append(COL_NUMERIC)
            continue

        # 5. Default to TEXT
        classifications.append(COL_TEXT)

    return classifications


# ===========================================================================
#  MODULE 6: TYPE APPLICATION
# ===========================================================================
def _to_amount(val):
    """Convert string to float for amount columns."""
    if val is None or str(val).strip() == "":
        return ""
    clean = str(val).strip().replace(",", "")
    if clean.endswith("-") and not clean.startswith("-"):
        clean = "-" + clean[:-1]
    try:
        return float(clean)
    except ValueError:
        return val


def _format_expiration(val):
    """Pad expiration value to 4 digits: '425' -> '0425'."""
    if val is None or str(val).strip() == "":
        return ""
    s = str(val).strip()
    # Handle float representation like '425.0'
    if "." in s:
        try:
            s = str(int(float(s)))
        except (ValueError, OverflowError):
            pass
    return s.zfill(4)


def to_number(val):
    """Convert string to int or float. Handles trailing minus and commas."""
    if val is None or str(val).strip() == "":
        return val
    clean = str(val).strip().replace(",", "")
    if clean.endswith("-") and not clean.startswith("-"):
        clean = "-" + clean[:-1]
    try:
        f = float(clean)
        if f == int(f) and "." not in clean:
            return int(f)
        return f
    except ValueError:
        return val


def apply_smart_types(rows, classifications):
    """Apply column classifications to all data rows."""
    cleaned = []
    for r_idx, row in enumerate(rows):
        new_row = []
        for c_idx, val in enumerate(row):
            if r_idx == 0:
                # Header -- always text
                new_row.append(val)
            elif c_idx < len(classifications):
                cls = classifications[c_idx]
                if cls == COL_ACCOUNT:
                    new_row.append(str(val).strip())
                elif cls == COL_AMOUNT:
                    new_row.append(_to_amount(val))
                elif cls == COL_EXPIRATION:
                    new_row.append(_format_expiration(val))
                elif cls == COL_NUMERIC:
                    new_row.append(to_number(val))
                else:
                    new_row.append(val)
            else:
                new_row.append(val)
        cleaned.append(new_row)
    return cleaned


# ===========================================================================
#  MODULE 7: DATA PROFILING (improved)
# ===========================================================================
def profile_data(rows, classifications):
    """Generate meaningful column profile with classification info."""
    if not rows:
        return []

    header = rows[0]
    data_rows = rows[1:]
    num_cols = len(header)
    total_records = len(data_rows)
    profiles = []

    for col_idx in range(num_cols):
        col_name = header[col_idx] if col_idx < len(header) else "Column_{}".format(col_idx + 1)
        cls = classifications[col_idx] if col_idx < len(classifications) else COL_TEXT

        values = []
        for row in data_rows:
            values.append(row[col_idx] if col_idx < len(row) else "")

        non_blank = [v for v in values if v is not None and str(v).strip() != ""]
        filled = len(non_blank)
        blank = total_records - filled
        completeness = round((filled / total_records) * 100, 1) if total_records > 0 else 0

        str_values = [str(v) for v in non_blank]
        unique_count = len(set(str_values))

        # Most common value
        most_common_val = ""
        most_common_freq = 0
        if str_values:
            counter = Counter(str_values)
            mc = counter.most_common(1)[0]
            most_common_val = mc[0]
            most_common_freq = mc[1]

        # Numeric stats
        min_val = ""
        max_val = ""
        avg_val = ""
        sum_val = ""

        if cls in (COL_NUMERIC, COL_AMOUNT):
            nums = []
            for v in non_blank:
                try:
                    n = float(str(v).replace(",", ""))
                    nums.append(n)
                except (ValueError, TypeError):
                    pass
            if nums:
                min_val = round(min(nums), 2)
                max_val = round(max(nums), 2)
                avg_val = round(sum(nums) / len(nums), 2)
                if cls == COL_AMOUNT:
                    sum_val = round(sum(nums), 2)

        profiles.append({
            "column_name": col_name,
            "classification": cls,
            "total_records": total_records,
            "filled": filled,
            "completeness": "{}%".format(completeness),
            "blank": blank,
            "unique": unique_count,
            "most_common": most_common_val,
            "frequency": most_common_freq,
            "min": min_val,
            "max": max_val,
            "avg": avg_val,
            "sum": sum_val,
        })
    return profiles


# ===========================================================================
#  MODULE 8: EXCEL WRITING
# ===========================================================================
def write_sheet(ws, rows, classifications=None):
    """Write rows with professional formatting."""
    if not rows:
        return

    for r_idx, row in enumerate(rows, start=1):
        for c_idx, val in enumerate(row, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            if r_idx == 1:
                cell.font = HEADER_FONT
                cell.fill = HEADER_FILL
                cell.alignment = HEADER_ALIGN
                cell.border = HEADER_BORDER
            else:
                cell.alignment = DATA_ALIGN
                # Apply number formats based on classification
                if classifications and c_idx - 1 < len(classifications):
                    cls = classifications[c_idx - 1]
                    if cls == COL_ACCOUNT:
                        cell.number_format = "@"
                    elif cls == COL_AMOUNT:
                        cell.number_format = "#,##0.00"
                    elif cls == COL_EXPIRATION:
                        cell.number_format = "@"
                    elif cls == COL_NUMERIC:
                        if isinstance(val, float):
                            cell.number_format = "#,##0.00"
                        elif isinstance(val, int):
                            cell.number_format = "#,##0"

    # Set header row height
    ws.row_dimensions[1].height = 30

    # Freeze panes
    ws.freeze_panes = "A2"

    # Auto-filter
    max_col = max(len(r) for r in rows) if rows else 1
    max_row = len(rows)
    ws.auto_filter.ref = "A1:{}{}".format(get_column_letter(max_col), max_row)

    # Auto-fit column widths
    for col_idx in range(1, max_col + 1):
        max_len = 0
        for row in rows:
            if col_idx - 1 < len(row):
                cell_val = row[col_idx - 1]
                cell_len = len(str(cell_val)) if cell_val is not None else 0
                if cell_len > max_len:
                    max_len = cell_len
        width = max(min(max_len + 3, 60), 10)
        ws.column_dimensions[get_column_letter(col_idx)].width = width


def write_profile_sheet(wb, sheet_name, profiles):
    """Add an improved data-profile sheet to the workbook."""
    ws = wb.create_sheet(title=sheet_name[:31])
    header = [
        "Column Name", "Classification", "Total Records", "Filled",
        "Completeness", "Blank", "Unique Values", "Most Common Value",
        "Frequency", "Min", "Max", "Average", "Sum (Amounts)",
    ]
    rows = [header]
    for p in profiles:
        rows.append([
            p["column_name"], p["classification"], p["total_records"],
            p["filled"], p["completeness"], p["blank"], p["unique"],
            p["most_common"], p["frequency"],
            p["min"], p["max"], p["avg"], p["sum"],
        ])
    write_sheet(ws, rows)
    for col_idx in range(1, len(header) + 1):
        ws.cell(row=1, column=col_idx).fill = PROFILE_HEADER_FILL


def write_report(rows, output_path, sheet_name, classifications=None,
                 profiles=None):
    """Write complete report .xlsx: data sheet + profile sheet."""
    wb = openpyxl.Workbook()
    ws = wb.active

    clean_name = sheet_name
    for ch in ["/", "\\", "?", "*", "[", "]", ":"]:
        clean_name = clean_name.replace(ch, "-")
    clean_name = clean_name[:31]
    ws.title = clean_name

    write_sheet(ws, rows, classifications)

    if profiles:
        pname = (clean_name[:22] + "_Profile"
                 if len(clean_name) > 22 else clean_name + "_Profile")
        write_profile_sheet(wb, pname[:31], profiles)

    wb.save(output_path)
    wb.close()


# ===========================================================================
#  MODULE 9: SUMMARY GENERATION
# ===========================================================================
def generate_summary(results, output_path):
    """Create Report_Summary.xlsx."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Processing Summary"

    header = [
        "Input File", "Output Name", "Row Count", "Column Count",
        "File Size (KB)", "Delimiter", "Status", "Timestamp",
    ]
    rows = [header]
    for r in results:
        delim = r.get("delimiter", "|")
        delim_display = "TAB" if delim == "\t" else delim
        rows.append([
            r.get("input_file", ""),
            r.get("output_name", ""),
            r.get("row_count", 0),
            r.get("col_count", 0),
            r.get("size_kb", 0),
            delim_display,
            r.get("status", ""),
            r.get("timestamp", ""),
        ])

    write_sheet(ws, rows)
    for col_idx in range(1, len(header) + 1):
        ws.cell(row=1, column=col_idx).fill = SUMMARY_HEADER_FILL

    wb.save(output_path)
    wb.close()


# ===========================================================================
#  MODULE 10: SINGLE FILE PROCESSOR
# ===========================================================================
def process_single_file(file_info, output_folder, output_mode="F",
                        tab_wb=None):
    """Process one report file end-to-end. Returns (result_dict, tab_wb)."""
    result = {
        "input_file": file_info["name"],
        "output_name": "",
        "row_count": 0,
        "col_count": 0,
        "size_kb": file_info["size_kb"],
        "delimiter": "",
        "status": "",
        "timestamp": datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
    }

    try:
        friendly = map_filename(file_info["name"])
        result["output_name"] = friendly + ".xlsx"

        delim = detect_delimiter(file_info["path"])
        result["delimiter"] = delim

        rows = parse_file(file_info["path"], delim)
        if not rows:
            result["status"] = "Empty File - Skipped"
            return result, tab_wb

        result["row_count"] = len(rows) - 1
        result["col_count"] = max(len(r) for r in rows)

        # Smart classification
        classifications = classify_columns(rows)
        rows = apply_smart_types(rows, classifications)
        profiles = profile_data(rows, classifications)

        sheet_name = friendly

        if output_mode == "F":
            out_path = os.path.join(output_folder, friendly + ".xlsx")
            write_report(rows, out_path, sheet_name, classifications,
                         profiles)
        else:
            if tab_wb is None:
                tab_wb = openpyxl.Workbook()
                tab_wb.remove(tab_wb.active)

            clean_name = sheet_name
            for ch in ["/", "\\", "?", "*", "[", "]", ":"]:
                clean_name = clean_name.replace(ch, "-")
            clean_name = clean_name[:31]

            ws = tab_wb.create_sheet(title=clean_name)
            write_sheet(ws, rows, classifications)

            if profiles:
                pname = (clean_name[:22] + "_Profile"
                         if len(clean_name) > 22
                         else clean_name + "_Profile")
                write_profile_sheet(tab_wb, pname[:31], profiles)

        result["status"] = "Success"

    except Exception as e:
        result["status"] = "Error: {}".format(str(e))

    return result, tab_wb


# ===========================================================================
#  MODULE 11: MAIN ORCHESTRATOR (AI Agent Ready)
# ===========================================================================
def run_report_tool(folder_path, output_mode="F", file_selection=None):
    """
    AI-Agent entry point. No user interaction.

    Args:
        folder_path:    Folder containing .txt / .txt.zip files.
        output_mode:    "F" for separate files, "T" for combined workbook.
        file_selection: List of filenames to process, or None for defaults.

    Returns:
        List of result dicts.
    """
    all_files = discover_files(folder_path)
    if not all_files:
        print("  No .txt or .txt.zip files found in: {}".format(folder_path))
        return []

    if file_selection is not None:
        sel_upper = [s.upper() for s in file_selection]
        selected = [f for f in all_files if f["name"].upper() in sel_upper]
    else:
        selected, _ = select_default_files(all_files)

    results = []
    tab_wb = None

    for num, finfo in enumerate(selected, start=1):
        friendly = map_filename(finfo["name"])
        print("")
        print("  [{}/{}] {} -> {}.xlsx".format(
            num, len(selected), finfo["name"], friendly))

        result, tab_wb = process_single_file(
            finfo, folder_path, output_mode, tab_wb)
        results.append(result)

        if result["status"] == "Success":
            print("         Rows: {} | Cols: {} | Delimiter: '{}' | OK".format(
                result["row_count"], result["col_count"],
                "TAB" if result["delimiter"] == "\t"
                else result["delimiter"]))
        else:
            print("         {}".format(result["status"]))

    if output_mode == "T" and tab_wb is not None:
        combined = os.path.join(folder_path, "Report_Output.xlsx")
        tab_wb.save(combined)
        tab_wb.close()
        print("")
        print("  Combined workbook -> {}".format(combined))

    summary_path = os.path.join(folder_path, "Report_Summary.xlsx")
    generate_summary(results, summary_path)
    print("  Summary -> Report_Summary.xlsx")

    return results


# ===========================================================================
#  INTERACTIVE MAIN
# ===========================================================================
def _ask_yn(prompt):
    """Ask a Y/N question. Returns True for Y, False for N."""
    while True:
        answer = input(prompt).strip().upper()
        if answer in ("Y", "YES"):
            return True
        elif answer in ("N", "NO", ""):
            return False
        else:
            print("    Please enter Y or N.")


def main():
    print("")
    print("=" * 70)
    print("  REPORT TOOL  -  Standalone Python Report Generator  v2")
    print("=" * 70)
    print("")

    # Step 1: Get folder path
    print("  Enter the folder path containing your report files")
    print("  (.txt or .txt.zip files):")
    print("")
    folder_path = input("  > ").strip()

    if (folder_path.startswith('"') and folder_path.endswith('"')) or \
       (folder_path.startswith("'") and folder_path.endswith("'")):
        folder_path = folder_path[1:-1]
    folder_path = folder_path.rstrip("\\/").strip()

    if not os.path.isdir(folder_path):
        print("")
        print("  ERROR: Folder not found: {}".format(folder_path))
        print("  Tips:")
        print("    - Copy-paste the path from File Explorer address bar")
        print("    - Special characters like & in folder names are OK")
        sys.exit(1)

    # Step 2: Discover files
    all_files = discover_files(folder_path)
    if not all_files:
        print("")
        print("  ERROR: No .txt or .txt.zip files found in:")
        print("  {}".format(folder_path))
        print("")
        print("  Files in this folder:")
        for f in sorted(os.listdir(folder_path))[:25]:
            print("    {}".format(f))
        sys.exit(1)

    # Step 3: Auto-select known reports
    selected_files, other_files = select_default_files(all_files)

    # Display selected files
    print("")
    print("-" * 70)
    print("  REPORT FILES TO PROCESS ({} of {} total):".format(
        len(selected_files), len(all_files)))
    print("")
    print("  {:<5} {:<25} {:<28} {:>10}".format(
        "#", "Input File", "Output Name", "Size (KB)"))
    print("  " + "-" * 70)
    for idx, finfo in enumerate(selected_files, start=1):
        friendly = map_filename(finfo["name"])
        print("  {:<5} {:<25} {:<28} {:>10}".format(
            idx, finfo["name"], friendly + ".xlsx", finfo["size_kb"]))

    if other_files:
        print("")
        print("  OTHER FILES (not selected by default):")
        for finfo in other_files:
            print("    - {}".format(finfo["name"]))

    # Step 4: Defaults
    print("")
    print("-" * 70)
    print("  DEFAULT SETTINGS:")
    print("    Output Mode     : F (each file saved as separate .xlsx)")
    print("    Files           : {} known report files".format(
        len(selected_files)))
    print("    Delimiter       : Auto-detected per file")
    print("    Column Types    : Smart detection (accounts as text,")
    print("                      amounts with 2 decimals, expiry padded)")
    print("    Formatting      : Centered cells, dark navy headers")
    print("    Add-ons         : Summary sheet + Data Profile per report")
    print("-" * 70)
    print("")

    customized = _ask_yn("  Do you want any customized changes? (Y/N): ")

    output_mode = "F"

    if customized:
        # Ask output mode
        print("")
        print("  OUTPUT MODE:")
        print("    F = Each report as a separate .xlsx file (default)")
        print("    T = All reports as tabs in one combined workbook")
        print("    Press Enter to keep default (F)")
        print("")
        mode_input = input(
            "    Choose F or T (or Enter to skip): ").strip().upper()
        while mode_input not in ("F", "T", ""):
            print("    Invalid. Please enter F or T only.")
            mode_input = input(
                "    Choose F or T (or Enter to skip): ").strip().upper()
        if mode_input in ("F", "T"):
            output_mode = mode_input
        print("    -> Output Mode: {} ({})".format(
            output_mode,
            "Separate Files" if output_mode == "F"
            else "Tabs in One Workbook"))

        # Ask file selection
        print("")
        has_changes = _ask_yn(
            "  Do you want to change which files to process? (Y/N): ")
        if has_changes:
            print("")
            print("  ALL AVAILABLE FILES:")
            combined_list = selected_files + other_files
            sel_names = set(f["name"] for f in selected_files)
            for idx, finfo in enumerate(combined_list, start=1):
                marker = " *" if finfo["name"] in sel_names else "  "
                print("    [{}]{} {} -> {}.xlsx".format(
                    idx, marker, finfo["name"],
                    map_filename(finfo["name"])))
            print("")
            print("  (* = currently selected)")
            print("  Enter the numbers of files to process")
            print("  (comma-separated), or Enter to keep current:")
            sel_input = input("    Process: ").strip()
            if sel_input:
                try:
                    sel_nums = [int(x.strip()) for x in sel_input.split(",")]
                    selected_files = [
                        combined_list[n - 1] for n in sel_nums
                        if 0 < n <= len(combined_list)]
                    print("")
                    print("  Updated ({} files):".format(
                        len(selected_files)))
                    for idx, finfo in enumerate(selected_files, start=1):
                        print("    [{}] {} -> {}.xlsx".format(
                            idx, finfo["name"],
                            map_filename(finfo["name"])))
                except (ValueError, IndexError):
                    print("    Invalid input. Keeping current selection.")

    # Step 5: Process
    print("")
    print("=" * 70)
    print("  PROCESSING {} file(s) in {} mode...".format(
        len(selected_files),
        "FILE (separate .xlsx)" if output_mode == "F"
        else "TAB (combined workbook)"))
    print("=" * 70)

    results = []
    tab_wb = None

    for num, finfo in enumerate(selected_files, start=1):
        friendly = map_filename(finfo["name"])
        print("")
        print("  [{}/{}] {}".format(
            num, len(selected_files), "-" * 55))
        print("         Input      : {}".format(finfo["name"]))
        print("         Output     : {}.xlsx".format(friendly))
        print("         Size       : {} KB".format(finfo["size_kb"]))

        result, tab_wb = process_single_file(
            finfo, folder_path, output_mode, tab_wb)
        results.append(result)

        if result["status"] == "Success":
            delim_d = ("TAB" if result["delimiter"] == "\t"
                       else result["delimiter"])
            print("         Delimiter  : '{}'  (auto-detected)".format(
                delim_d))
            print("         Rows       : {:,}".format(result["row_count"]))
            print("         Columns    : {}".format(result["col_count"]))
            if output_mode == "F":
                print("         Result     : SAVED -> {}.xlsx".format(
                    friendly))
            else:
                print("         Result     : ADDED as tab '{}'".format(
                    friendly))
        else:
            print("         Result     : {}".format(result["status"]))

    if output_mode == "T" and tab_wb is not None:
        combined_path = os.path.join(folder_path, "Report_Output.xlsx")
        tab_wb.save(combined_path)
        tab_wb.close()
        print("")
        print("  Combined workbook saved -> Report_Output.xlsx")

    summary_path = os.path.join(folder_path, "Report_Summary.xlsx")
    generate_summary(results, summary_path)

    # Final summary
    print("")
    print("=" * 70)
    print("  PROCESSING COMPLETE")
    print("=" * 70)
    print("")
    ok = sum(1 for r in results if r["status"] == "Success")
    err = sum(1 for r in results if r["status"] != "Success")
    total_rows = sum(
        r["row_count"] for r in results if r["status"] == "Success")

    print("  Files Processed : {}".format(len(results)))
    print("  Successful      : {}".format(ok))
    if err > 0:
        print("  Errors          : {}".format(err))
    print("  Total Data Rows : {:,}".format(total_rows))
    print("  Output Mode     : {}".format(
        "Separate Files" if output_mode == "F" else "Combined Workbook"))
    print("")
    print("  Output Location : {}".format(folder_path))
    print("")
    print("  Files Created:")
    if output_mode == "F":
        for r in results:
            if r["status"] == "Success":
                print("    -> {}".format(r["output_name"]))
    else:
        print("    -> Report_Output.xlsx ({} tabs)".format(ok))
    print("    -> Report_Summary.xlsx")
    print("")
    print("=" * 70)
    print("")


if __name__ == "__main__":
    main()
